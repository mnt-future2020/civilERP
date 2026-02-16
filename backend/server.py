from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import httpx
import hashlib
import base64
import json
import qrcode
import shutil
import cloudinary
import cloudinary.uploader
from io import BytesIO
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import jwt
from passlib.context import CryptContext
from cryptography.fernet import Fernet
from openai import AsyncOpenAI
import xmltodict

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT and Password Config
JWT_SECRET = os.environ.get('JWT_SECRET', 'civil_erp_secret_key')
JWT_ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_HOURS = 24

# Encryption key for storing credentials
FERNET_KEY = os.environ.get('FERNET_KEY', Fernet.generate_key().decode())
fernet = Fernet(FERNET_KEY.encode() if isinstance(FERNET_KEY, str) else FERNET_KEY)

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

# Create the main app
app = FastAPI(title="Civil Construction ERP API")
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== MODELS ====================

class UserRole:
    ADMIN = "admin"
    SITE_ENGINEER = "site_engineer"
    FINANCE = "finance"
    PROCUREMENT = "procurement"

# ==================== RBAC MODELS ====================

# Available modules for permission assignment
AVAILABLE_MODULES = [
    "dashboard",
    "projects", 
    "financial",
    "procurement",
    "hrms",
    "compliance",
    "einvoicing",
    "reports",
    "ai_assistant",
    "settings",
    "admin"
]

# Available actions per module
AVAILABLE_ACTIONS = ["view", "create", "edit", "delete"]

class ModulePermission(BaseModel):
    module: str
    view: bool = False
    create: bool = False
    edit: bool = False
    delete: bool = False

class RoleCreate(BaseModel):
    name: str
    description: Optional[str] = None
    permissions: List[ModulePermission] = []

class RoleUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    permissions: Optional[List[ModulePermission]] = None
    is_active: Optional[bool] = None

class Role(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    permissions: List[Dict] = []
    is_system_role: bool = False  # System roles cannot be deleted
    is_active: bool = True
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class UserBase(BaseModel):
    email: EmailStr
    name: str
    role: str = UserRole.SITE_ENGINEER  # Legacy field for backwards compatibility
    role_id: Optional[str] = None  # New RBAC role reference
    phone: Optional[str] = None
    department: Optional[str] = None

class UserCreate(UserBase):
    password: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserRoleAssignment(BaseModel):
    user_id: str
    role_id: str

class User(UserBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    is_active: bool = True

class UserWithPermissions(User):
    permissions: Dict[str, Dict[str, bool]] = {}
    role_name: Optional[str] = None

class Token(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserWithPermissions

# Project Models
class ProjectStatus:
    PLANNING = "planning"
    IN_PROGRESS = "in_progress"
    ON_HOLD = "on_hold"
    COMPLETED = "completed"

class ProjectCreate(BaseModel):
    name: str
    code: str
    description: Optional[str] = None
    client_name: str
    location: str
    start_date: str
    expected_end_date: str
    budget: float
    site_engineer_id: Optional[str] = None

class Project(ProjectCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    status: str = ProjectStatus.PLANNING
    actual_cost: float = 0.0
    progress_percentage: float = 0.0
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    created_by: Optional[str] = None

# Task Models (WBS)
class TaskCreate(BaseModel):
    project_id: str
    name: str
    description: Optional[str] = None
    parent_task_id: Optional[str] = None
    start_date: str
    end_date: str
    estimated_cost: float = 0.0
    assigned_to: Optional[str] = None

class Task(TaskCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    status: str = "pending"
    actual_cost: float = 0.0
    progress: float = 0.0
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Daily Progress Report
class DPRCreate(BaseModel):
    project_id: str
    date: str
    weather: Optional[str] = None
    labor_count: int = 0
    work_done: str
    materials_used: Optional[str] = None
    issues: Optional[str] = None
    notes: Optional[str] = None

class DPR(DPRCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_by: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Financial Models - CVR
class CVRCreate(BaseModel):
    project_id: str
    period_start: str
    period_end: str
    contracted_value: float
    work_done_value: float
    billed_value: float
    received_value: float
    retention_held: float = 0.0
    notes: Optional[str] = None

class CVR(CVRCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    variance: float = 0.0
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Billing
class BillingCreate(BaseModel):
    project_id: str
    bill_number: str
    bill_date: str
    description: str
    amount: float
    gst_rate: float = 18.0
    bill_type: str = "running"  # running, final, advance

class Billing(BillingCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    gst_amount: float = 0.0
    total_amount: float = 0.0
    status: str = "pending"
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Vendor Models
class VendorCreate(BaseModel):
    name: str
    gstin: Optional[str] = None
    pan: Optional[str] = None
    address: str
    city: str
    state: str = "Tamil Nadu"
    pincode: str
    contact_person: str
    phone: str
    email: EmailStr
    category: str  # material, labor, equipment, subcontractor

class Vendor(VendorCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    is_active: bool = True
    rating: float = 0.0
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Purchase Order Models
class POItemCreate(BaseModel):
    description: str
    unit: str
    quantity: float
    rate: float

class PurchaseOrderCreate(BaseModel):
    project_id: str
    vendor_id: str
    po_date: str
    delivery_date: str
    items: List[POItemCreate]
    terms: Optional[str] = None

class PurchaseOrder(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    po_number: str = ""
    project_id: str
    vendor_id: str
    po_date: str
    delivery_date: str
    items: List[Dict]
    terms: Optional[str] = None
    subtotal: float = 0.0
    gst_amount: float = 0.0
    total: float = 0.0
    status: str = "pending"
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# GRN Models
class GRNItemCreate(BaseModel):
    po_item_index: int
    received_quantity: float
    remarks: Optional[str] = None

class GRNCreate(BaseModel):
    po_id: str
    grn_date: str
    items: List[GRNItemCreate]
    notes: Optional[str] = None

class GRN(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    grn_number: str = ""
    po_id: str
    grn_date: str
    items: List[Dict]
    notes: Optional[str] = None
    status: str = "received"
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Employee Models
class EmployeeCreate(BaseModel):
    name: str
    employee_code: str
    designation: str
    department: str
    phone: str
    email: EmailStr
    date_of_joining: str
    basic_salary: float
    hra: float = 0.0
    pf_number: Optional[str] = None
    esi_number: Optional[str] = None
    bank_account: Optional[str] = None
    bank_name: Optional[str] = None
    ifsc: Optional[str] = None
    user_id: Optional[str] = None  # Link to auth user account
    create_user_account: bool = False  # Flag to create user account on employee creation

class Employee(EmployeeCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    is_active: bool = True
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Attendance Models
class AttendanceCreate(BaseModel):
    employee_id: str
    project_id: str
    date: str
    check_in: Optional[str] = None
    check_out: Optional[str] = None
    status: str = "present"  # present, absent, half_day, leave
    overtime_hours: float = 0.0

class Attendance(AttendanceCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Payroll Models
class PayrollCreate(BaseModel):
    employee_id: str
    month: str  # YYYY-MM
    basic_salary: float
    hra: float = 0.0
    overtime_pay: float = 0.0
    other_allowances: float = 0.0
    pf_deduction: float = 0.0
    esi_deduction: float = 0.0
    tds: float = 0.0
    other_deductions: float = 0.0

class Payroll(PayrollCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    gross_salary: float = 0.0
    total_deductions: float = 0.0
    net_salary: float = 0.0
    status: str = "pending"
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# GST Models
class GSTReturnCreate(BaseModel):
    return_type: str  # GSTR-1, GSTR-3B
    period: str  # YYYY-MM
    total_outward_supplies: float = 0.0
    total_inward_supplies: float = 0.0
    cgst: float = 0.0
    sgst: float = 0.0
    igst: float = 0.0
    itc_claimed: float = 0.0

class GSTReturn(GSTReturnCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tax_payable: float = 0.0
    status: str = "draft"
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# RERA Models
class RERAProjectCreate(BaseModel):
    project_id: str
    rera_number: str
    registration_date: str
    validity_date: str
    escrow_bank: str
    escrow_account: str
    total_units: int
    sold_units: int = 0

class RERAProject(RERAProjectCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    compliance_status: str = "compliant"
    last_quarterly_update: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# ==================== GST E-INVOICE MODELS ====================

class EInvoiceItemCreate(BaseModel):
    sl_no: int
    item_description: str
    hsn_code: str
    quantity: float
    unit: str = "NOS"
    unit_price: float
    discount: float = 0.0
    taxable_value: float
    gst_rate: float = 18.0
    cgst_amount: float = 0.0
    sgst_amount: float = 0.0
    igst_amount: float = 0.0
    cess_amount: float = 0.0
    total_item_value: float = 0.0

class EInvoiceCreate(BaseModel):
    billing_id: Optional[str] = None
    # Transaction Details
    supply_type: str = "B2B"  # B2B, B2C, SEZWP, SEZWOP, EXPWP, EXPWOP, DEXP
    document_type: str = "INV"  # INV, CRN, DBN
    document_number: str
    document_date: str
    
    # Seller Details
    seller_gstin: str
    seller_legal_name: str
    seller_trade_name: Optional[str] = None
    seller_address: str
    seller_location: str
    seller_pincode: str
    seller_state_code: str = "33"  # Tamil Nadu
    
    # Buyer Details
    buyer_gstin: str
    buyer_legal_name: str
    buyer_trade_name: Optional[str] = None
    buyer_address: str
    buyer_location: str
    buyer_pincode: str
    buyer_state_code: str = "33"
    buyer_pos: str = "33"  # Place of Supply
    
    # Dispatch Details (Optional)
    dispatch_from_name: Optional[str] = None
    dispatch_from_address: Optional[str] = None
    dispatch_from_location: Optional[str] = None
    dispatch_from_pincode: Optional[str] = None
    dispatch_from_state_code: Optional[str] = None
    
    # Ship To Details (Optional)
    ship_to_gstin: Optional[str] = None
    ship_to_legal_name: Optional[str] = None
    ship_to_address: Optional[str] = None
    ship_to_location: Optional[str] = None
    ship_to_pincode: Optional[str] = None
    ship_to_state_code: Optional[str] = None
    
    # Items
    items: List[EInvoiceItemCreate]
    
    # Value Details
    total_taxable_value: float
    total_cgst: float = 0.0
    total_sgst: float = 0.0
    total_igst: float = 0.0
    total_cess: float = 0.0
    total_discount: float = 0.0
    other_charges: float = 0.0
    round_off: float = 0.0
    total_invoice_value: float
    
    # Payment Details
    payment_mode: str = "CREDIT"  # CASH, CREDIT, DIRECT_TRANSFER
    payment_terms: Optional[str] = None
    
    # E-Way Bill Details (Optional)
    transporter_id: Optional[str] = None
    transporter_name: Optional[str] = None
    transport_mode: Optional[str] = None  # 1-Road, 2-Rail, 3-Air, 4-Ship
    transport_distance: Optional[int] = None
    vehicle_number: Optional[str] = None
    vehicle_type: Optional[str] = None

class EInvoice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    billing_id: Optional[str] = None
    document_number: str
    document_date: str
    document_type: str
    supply_type: str
    seller_gstin: str
    seller_legal_name: str
    buyer_gstin: str
    buyer_legal_name: str
    total_taxable_value: float
    total_cgst: float
    total_sgst: float
    total_igst: float
    total_invoice_value: float
    items: List[Dict]
    
    # IRN Details (populated after NIC submission)
    irn: Optional[str] = None
    ack_number: Optional[str] = None
    ack_date: Optional[str] = None
    signed_invoice: Optional[str] = None
    signed_qr_code: Optional[str] = None
    qr_code_image: Optional[str] = None  # Base64 encoded
    
    # E-Way Bill Details
    eway_bill_number: Optional[str] = None
    eway_bill_date: Optional[str] = None
    eway_bill_valid_till: Optional[str] = None
    
    # Status
    status: str = "draft"  # draft, submitted, irn_generated, cancelled, rejected
    nic_response: Optional[Dict] = None
    error_details: Optional[str] = None
    
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: Optional[str] = None

class GSTINVerification(BaseModel):
    gstin: str
    legal_name: Optional[str] = None
    trade_name: Optional[str] = None
    registration_date: Optional[str] = None
    constitution_of_business: Optional[str] = None
    taxpayer_type: Optional[str] = None
    gstin_status: Optional[str] = None
    state_jurisdiction: Optional[str] = None
    centre_jurisdiction: Optional[str] = None
    address: Optional[str] = None
    is_valid: bool = False
    verified_at: Optional[str] = None

class EWayBillCreate(BaseModel):
    einvoice_id: str
    transporter_id: str
    transporter_name: str
    transport_mode: str  # 1-Road, 2-Rail, 3-Air, 4-Ship
    transport_distance: int
    vehicle_number: Optional[str] = None
    vehicle_type: Optional[str] = None  # R-Regular, O-Over Dimensional

class EWayBill(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    einvoice_id: str
    eway_bill_number: str
    eway_bill_date: str
    valid_till: str
    transporter_id: str
    transporter_name: str
    transport_mode: str
    vehicle_number: Optional[str] = None
    status: str = "active"  # active, cancelled, expired
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# AI Request Model
class AIRequest(BaseModel):
    query: str
    context: Optional[Dict] = None

# ==================== AUTH HELPERS ====================

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(hours=ACCESS_TOKEN_EXPIRE_HOURS)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_user_permissions(user_doc: dict) -> Dict[str, Dict[str, bool]]:
    """Get user permissions based on their assigned role"""
    permissions = {}
    
    # Initialize all modules with no permissions
    for module in AVAILABLE_MODULES:
        permissions[module] = {"view": False, "create": False, "edit": False, "delete": False}
    
    # Check for RBAC role_id first
    if user_doc.get("role_id"):
        role_doc = await db.roles.find_one({"id": user_doc["role_id"], "is_active": True}, {"_id": 0})
        if role_doc:
            for perm in role_doc.get("permissions", []):
                module = perm.get("module")
                if module in permissions:
                    permissions[module] = {
                        "view": perm.get("view", False),
                        "create": perm.get("create", False),
                        "edit": perm.get("edit", False),
                        "delete": perm.get("delete", False)
                    }
            return permissions
    
    # Fallback to legacy role-based permissions
    legacy_role = user_doc.get("role", "site_engineer")
    
    # Admin has full access
    if legacy_role == "admin":
        for module in AVAILABLE_MODULES:
            permissions[module] = {"view": True, "create": True, "edit": True, "delete": True}
    elif legacy_role == "site_engineer":
        permissions["dashboard"] = {"view": True, "create": False, "edit": False, "delete": False}
        permissions["projects"] = {"view": True, "create": True, "edit": True, "delete": False}
        permissions["reports"] = {"view": True, "create": False, "edit": False, "delete": False}
        permissions["ai_assistant"] = {"view": True, "create": True, "edit": False, "delete": False}
    elif legacy_role == "finance":
        permissions["dashboard"] = {"view": True, "create": False, "edit": False, "delete": False}
        permissions["projects"] = {"view": True, "create": False, "edit": False, "delete": False}
        permissions["financial"] = {"view": True, "create": True, "edit": True, "delete": True}
        permissions["compliance"] = {"view": True, "create": True, "edit": True, "delete": True}
        permissions["einvoicing"] = {"view": True, "create": True, "edit": True, "delete": True}
        permissions["reports"] = {"view": True, "create": True, "edit": False, "delete": False}
    elif legacy_role == "procurement":
        permissions["dashboard"] = {"view": True, "create": False, "edit": False, "delete": False}
        permissions["projects"] = {"view": True, "create": False, "edit": False, "delete": False}
        permissions["procurement"] = {"view": True, "create": True, "edit": True, "delete": True}
        permissions["reports"] = {"view": True, "create": False, "edit": False, "delete": False}
    
    return permissions

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> User:
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        user_doc = await db.users.find_one({"id": user_id}, {"_id": 0})
        if user_doc is None:
            raise HTTPException(status_code=401, detail="User not found")
        return User(**user_doc)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def get_current_user_with_permissions(credentials: HTTPAuthorizationCredentials = Depends(security)) -> UserWithPermissions:
    """Get current user with their permissions"""
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        user_doc = await db.users.find_one({"id": user_id}, {"_id": 0})
        if user_doc is None:
            raise HTTPException(status_code=401, detail="User not found")
        
        # Get permissions
        permissions = await get_user_permissions(user_doc)
        
        # Get role name if using RBAC
        role_name = None
        if user_doc.get("role_id"):
            role_doc = await db.roles.find_one({"id": user_doc["role_id"]}, {"_id": 0, "name": 1})
            if role_doc:
                role_name = role_doc.get("name")
        
        user_data = {k: v for k, v in user_doc.items() if k not in ['_id', 'password']}
        return UserWithPermissions(**user_data, permissions=permissions, role_name=role_name)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

def check_role(allowed_roles: List[str]):
    """Legacy role check decorator"""
    async def role_checker(current_user: User = Depends(get_current_user)) -> User:
        if current_user.role not in allowed_roles:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return current_user
    return role_checker

def require_permission(module: str, action: str):
    """RBAC permission check decorator"""
    async def permission_checker(credentials: HTTPAuthorizationCredentials = Depends(security)) -> UserWithPermissions:
        user = await get_current_user_with_permissions(credentials)
        
        # Check if user has required permission
        if module not in user.permissions:
            raise HTTPException(status_code=403, detail=f"Access to {module} denied")
        
        if not user.permissions[module].get(action, False):
            raise HTTPException(status_code=403, detail=f"Insufficient permissions: {action} on {module} denied")
        
        return user
    return permission_checker

def require_admin():
    """Decorator to require admin role"""
    async def admin_checker(current_user: User = Depends(get_current_user)) -> User:
        if current_user.role != "admin":
            raise HTTPException(status_code=403, detail="Admin access required")
        return current_user
    return admin_checker

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/register", response_model=Token)
async def register(user_data: UserCreate):
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_dict = user_data.model_dump()
    user_dict['password'] = get_password_hash(user_dict['password'])
    user_obj = User(**{k: v for k, v in user_dict.items() if k != 'password'})
    
    doc = {**user_obj.model_dump(), "password": user_dict['password']}
    await db.users.insert_one(doc)
    
    # Get user permissions for token response
    permissions = await get_user_permissions(doc)
    user_with_perms = UserWithPermissions(**user_obj.model_dump(), permissions=permissions)
    
    access_token = create_access_token({"sub": user_obj.id, "role": user_obj.role})
    return Token(access_token=access_token, user=user_with_perms)

@api_router.post("/auth/login", response_model=Token)
async def login(credentials: UserLogin):
    user_doc = await db.users.find_one({"email": credentials.email})
    if not user_doc or not verify_password(credentials.password, user_doc['password']):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Get permissions
    permissions = await get_user_permissions(user_doc)
    
    # Get role name if using RBAC
    role_name = None
    if user_doc.get("role_id"):
        role_doc = await db.roles.find_one({"id": user_doc["role_id"]}, {"_id": 0, "name": 1})
        if role_doc:
            role_name = role_doc.get("name")
    
    user_data = {k: v for k, v in user_doc.items() if k not in ['_id', 'password']}
    user_with_perms = UserWithPermissions(**user_data, permissions=permissions, role_name=role_name)
    
    access_token = create_access_token({"sub": user_with_perms.id, "role": user_with_perms.role})
    return Token(access_token=access_token, user=user_with_perms)

@api_router.get("/auth/me", response_model=UserWithPermissions)
async def get_me(current_user: UserWithPermissions = Depends(get_current_user_with_permissions)):
    return current_user

# ==================== RBAC ROUTES ====================

# Get available modules and actions for admin UI
@api_router.get("/rbac/modules")
async def get_available_modules(current_user: User = Depends(require_admin())):
    return {
        "modules": AVAILABLE_MODULES,
        "actions": AVAILABLE_ACTIONS
    }

# Roles CRUD
@api_router.post("/rbac/roles", response_model=Role)
async def create_role(role_data: RoleCreate, current_user: User = Depends(require_admin())):
    # Check if role name already exists
    existing = await db.roles.find_one({"name": role_data.name.strip()})
    if existing:
        raise HTTPException(status_code=400, detail="Role with this name already exists")
    
    # Validate permissions
    permissions_list = []
    for perm in role_data.permissions:
        if perm.module not in AVAILABLE_MODULES:
            raise HTTPException(status_code=400, detail=f"Invalid module: {perm.module}")
        permissions_list.append(perm.model_dump())
    
    role = Role(
        name=role_data.name.strip(),
        description=role_data.description,
        permissions=permissions_list,
        is_system_role=False
    )
    
    await db.roles.insert_one(role.model_dump())
    return role

@api_router.get("/rbac/roles", response_model=List[Role])
async def list_roles(
    include_inactive: bool = False,
    current_user: User = Depends(require_admin())
):
    query = {} if include_inactive else {"is_active": True}
    roles = await db.roles.find(query, {"_id": 0}).to_list(1000)
    return [Role(**r) for r in roles]

@api_router.get("/rbac/roles/{role_id}", response_model=Role)
async def get_role(role_id: str, current_user: User = Depends(require_admin())):
    role = await db.roles.find_one({"id": role_id}, {"_id": 0})
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    return Role(**role)

@api_router.put("/rbac/roles/{role_id}", response_model=Role)
async def update_role(role_id: str, role_data: RoleUpdate, current_user: User = Depends(require_admin())):
    role = await db.roles.find_one({"id": role_id}, {"_id": 0})
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    
    # Don't allow editing system roles completely (only permissions)
    if role.get("is_system_role") and role_data.name and role_data.name != role.get("name"):
        raise HTTPException(status_code=400, detail="Cannot rename system roles")
    
    update_data = {}
    if role_data.name is not None:
        # Check name uniqueness
        existing = await db.roles.find_one({"name": role_data.name.strip(), "id": {"$ne": role_id}})
        if existing:
            raise HTTPException(status_code=400, detail="Role with this name already exists")
        update_data["name"] = role_data.name.strip()
    
    if role_data.description is not None:
        update_data["description"] = role_data.description
    
    if role_data.permissions is not None:
        permissions_list = []
        for perm in role_data.permissions:
            if perm.module not in AVAILABLE_MODULES:
                raise HTTPException(status_code=400, detail=f"Invalid module: {perm.module}")
            permissions_list.append(perm.model_dump())
        update_data["permissions"] = permissions_list
    
    if role_data.is_active is not None:
        if role.get("is_system_role") and not role_data.is_active:
            raise HTTPException(status_code=400, detail="Cannot deactivate system roles")
        update_data["is_active"] = role_data.is_active
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.roles.update_one({"id": role_id}, {"$set": update_data})
    
    updated_role = await db.roles.find_one({"id": role_id}, {"_id": 0})
    return Role(**updated_role)

@api_router.delete("/rbac/roles/{role_id}")
async def delete_role(role_id: str, current_user: User = Depends(require_admin())):
    role = await db.roles.find_one({"id": role_id}, {"_id": 0})
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    
    if role.get("is_system_role"):
        raise HTTPException(status_code=400, detail="Cannot delete system roles")
    
    # Check if any users have this role
    users_with_role = await db.users.count_documents({"role_id": role_id})
    if users_with_role > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete role - {users_with_role} user(s) are assigned to it")
    
    await db.roles.delete_one({"id": role_id})
    return {"message": "Role deleted successfully"}

# User Role Assignment
@api_router.post("/rbac/assign-role")
async def assign_role_to_user(assignment: UserRoleAssignment, current_user: User = Depends(require_admin())):
    # Verify user exists
    user = await db.users.find_one({"id": assignment.user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Verify role exists and is active
    role = await db.roles.find_one({"id": assignment.role_id, "is_active": True}, {"_id": 0})
    if not role:
        raise HTTPException(status_code=404, detail="Role not found or inactive")
    
    # Update user with new role
    await db.users.update_one(
        {"id": assignment.user_id},
        {"$set": {"role_id": assignment.role_id}}
    )
    
    return {"message": f"Role '{role['name']}' assigned to user successfully"}

@api_router.delete("/rbac/users/{user_id}/role")
async def remove_role_from_user(user_id: str, current_user: User = Depends(require_admin())):
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    await db.users.update_one(
        {"id": user_id},
        {"$unset": {"role_id": ""}}
    )
    
    return {"message": "Role removed from user, falling back to legacy role permissions"}

# Get users with their roles for admin management
@api_router.get("/rbac/users")
async def list_users_with_roles(current_user: User = Depends(require_admin())):
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    roles = await db.roles.find({}, {"_id": 0}).to_list(1000)
    roles_map = {r["id"]: r["name"] for r in roles}
    
    result = []
    for u in users:
        user_data = {
            "id": u.get("id"),
            "email": u.get("email"),
            "name": u.get("name"),
            "role": u.get("role"),  # Legacy role
            "role_id": u.get("role_id"),
            "role_name": roles_map.get(u.get("role_id"), None),
            "department": u.get("department"),
            "is_active": u.get("is_active", True),
            "created_at": u.get("created_at")
        }
        result.append(user_data)
    
    return result

# Get role statistics for admin dashboard
@api_router.get("/rbac/stats")
async def get_rbac_stats(current_user: User = Depends(require_admin())):
    total_roles = await db.roles.count_documents({})
    active_roles = await db.roles.count_documents({"is_active": True})
    total_users = await db.users.count_documents({})
    users_with_rbac_role = await db.users.count_documents({"role_id": {"$exists": True, "$ne": None}})
    
    # Get user count per role
    pipeline = [
        {"$match": {"role_id": {"$exists": True, "$ne": None}}},
        {"$group": {"_id": "$role_id", "count": {"$sum": 1}}}
    ]
    role_counts = await db.users.aggregate(pipeline).to_list(1000)
    
    roles = await db.roles.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    roles_map = {r["id"]: r["name"] for r in roles}
    
    users_per_role = {roles_map.get(rc["_id"], "Unknown"): rc["count"] for rc in role_counts}
    
    return {
        "total_roles": total_roles,
        "active_roles": active_roles,
        "total_users": total_users,
        "users_with_rbac_role": users_with_rbac_role,
        "users_with_legacy_role": total_users - users_with_rbac_role,
        "users_per_role": users_per_role
    }

# Initialize default system roles
@api_router.post("/rbac/init")
async def initialize_system_roles(current_user: User = Depends(require_admin())):
    """Initialize default system roles if they don't exist"""
    
    default_roles = [
        {
            "name": "Administrator",
            "description": "Full system access",
            "is_system_role": True,
            "permissions": [{"module": m, "view": True, "create": True, "edit": True, "delete": True} for m in AVAILABLE_MODULES]
        },
        {
            "name": "HR Manager",
            "description": "Human Resources management",
            "is_system_role": False,
            "permissions": [
                {"module": "dashboard", "view": True, "create": False, "edit": False, "delete": False},
                {"module": "hrms", "view": True, "create": True, "edit": True, "delete": True},
                {"module": "reports", "view": True, "create": True, "edit": False, "delete": False},
            ]
        },
        {
            "name": "Project Manager",
            "description": "Project management and oversight",
            "is_system_role": False,
            "permissions": [
                {"module": "dashboard", "view": True, "create": False, "edit": False, "delete": False},
                {"module": "projects", "view": True, "create": True, "edit": True, "delete": False},
                {"module": "procurement", "view": True, "create": False, "edit": False, "delete": False},
                {"module": "reports", "view": True, "create": True, "edit": False, "delete": False},
                {"module": "ai_assistant", "view": True, "create": True, "edit": False, "delete": False},
            ]
        },
        {
            "name": "Accountant",
            "description": "Financial operations and compliance",
            "is_system_role": False,
            "permissions": [
                {"module": "dashboard", "view": True, "create": False, "edit": False, "delete": False},
                {"module": "financial", "view": True, "create": True, "edit": True, "delete": False},
                {"module": "compliance", "view": True, "create": True, "edit": True, "delete": False},
                {"module": "einvoicing", "view": True, "create": True, "edit": True, "delete": False},
                {"module": "reports", "view": True, "create": True, "edit": False, "delete": False},
            ]
        },
        {
            "name": "Site Engineer",
            "description": "On-site project execution",
            "is_system_role": False,
            "permissions": [
                {"module": "dashboard", "view": True, "create": False, "edit": False, "delete": False},
                {"module": "projects", "view": True, "create": True, "edit": True, "delete": False},
                {"module": "reports", "view": True, "create": False, "edit": False, "delete": False},
                {"module": "ai_assistant", "view": True, "create": True, "edit": False, "delete": False},
            ]
        },
        {
            "name": "Procurement Officer",
            "description": "Procurement and vendor management",
            "is_system_role": False,
            "permissions": [
                {"module": "dashboard", "view": True, "create": False, "edit": False, "delete": False},
                {"module": "procurement", "view": True, "create": True, "edit": True, "delete": True},
                {"module": "projects", "view": True, "create": False, "edit": False, "delete": False},
                {"module": "reports", "view": True, "create": False, "edit": False, "delete": False},
            ]
        },
    ]
    
    created_count = 0
    for role_data in default_roles:
        existing = await db.roles.find_one({"name": role_data["name"]})
        if not existing:
            role = Role(**role_data)
            await db.roles.insert_one(role.model_dump())
            created_count += 1
    
    return {"message": f"Initialized {created_count} system roles", "total_roles": len(default_roles)}

# ==================== DASHBOARD ====================

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: User = Depends(get_current_user)):
    total_projects = await db.projects.count_documents({})
    active_projects = await db.projects.count_documents({"status": "in_progress"})
    total_vendors = await db.vendors.count_documents({"is_active": True})
    total_employees = await db.employees.count_documents({"is_active": True})
    
    # Financial stats
    projects = await db.projects.find({}, {"_id": 0, "budget": 1, "actual_cost": 1}).to_list(1000)
    total_budget = sum(p.get('budget', 0) for p in projects)
    total_spent = sum(p.get('actual_cost', 0) for p in projects)
    
    # Recent POs
    pending_pos = await db.purchase_orders.count_documents({"status": "pending"})
    
    # Attendance today
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    present_today = await db.attendance.count_documents({"date": today, "status": "present"})
    
    return {
        "total_projects": total_projects,
        "active_projects": active_projects,
        "total_vendors": total_vendors,
        "total_employees": total_employees,
        "total_budget": total_budget,
        "total_spent": total_spent,
        "budget_utilization": (total_spent / total_budget * 100) if total_budget > 0 else 0,
        "pending_pos": pending_pos,
        "present_today": present_today,
        "spi": 0.97,  # Schedule Performance Index - would be calculated from tasks
        "cost_variance": total_budget - total_spent,
        "safety_incidents": 0,
        "equipment_utilization": 87.5
    }

@api_router.get("/dashboard/chart-data")
async def get_chart_data(current_user: User = Depends(get_current_user)):
    # Monthly cost data for charts
    projects = await db.projects.find({}, {"_id": 0}).to_list(100)
    
    # Simulate monthly data
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]
    budget_data = [120, 150, 180, 200, 220, 250]
    actual_data = [115, 155, 170, 210, 215, 240]
    
    return {
        "monthly_cost": {
            "labels": months,
            "budget": budget_data,
            "actual": actual_data
        },
        "project_status": {
            "planning": await db.projects.count_documents({"status": "planning"}),
            "in_progress": await db.projects.count_documents({"status": "in_progress"}),
            "on_hold": await db.projects.count_documents({"status": "on_hold"}),
            "completed": await db.projects.count_documents({"status": "completed"})
        },
        "expense_breakdown": {
            "materials": 45,
            "labor": 30,
            "equipment": 15,
            "overhead": 10
        }
    }

# ==================== PROJECT ROUTES ====================

@api_router.post("/projects", response_model=Project)
async def create_project(project_data: ProjectCreate, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.SITE_ENGINEER]))):
    project = Project(**project_data.model_dump(), created_by=current_user.id)
    doc = project.model_dump()
    await db.projects.insert_one(doc)
    return project

@api_router.get("/projects", response_model=List[Project])
async def get_projects(current_user: User = Depends(get_current_user)):
    projects = await db.projects.find({}, {"_id": 0}).to_list(1000)
    return projects

@api_router.get("/projects/{project_id}", response_model=Project)
async def get_project(project_id: str, current_user: User = Depends(get_current_user)):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return Project(**project)

@api_router.put("/projects/{project_id}", response_model=Project)
async def update_project(project_id: str, project_data: ProjectCreate, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.SITE_ENGINEER]))):
    existing = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Project not found")
    
    update_dict = project_data.model_dump()
    await db.projects.update_one({"id": project_id}, {"$set": update_dict})
    updated = await db.projects.find_one({"id": project_id}, {"_id": 0})
    return Project(**updated)

@api_router.delete("/projects/{project_id}")
async def delete_project(project_id: str, current_user: User = Depends(check_role([UserRole.ADMIN]))):
    result = await db.projects.delete_one({"id": project_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    return {"message": "Project deleted"}

class ProjectStatusUpdate(BaseModel):
    status: str

class ProjectProgressUpdate(BaseModel):
    progress_percentage: float
    actual_cost: Optional[float] = None

@api_router.patch("/projects/{project_id}/status")
async def update_project_status(project_id: str, data: ProjectStatusUpdate, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.SITE_ENGINEER]))):
    existing = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Project not found")
    await db.projects.update_one({"id": project_id}, {"$set": {"status": data.status}})
    updated = await db.projects.find_one({"id": project_id}, {"_id": 0})
    return updated

@api_router.patch("/projects/{project_id}/progress")
async def update_project_progress(project_id: str, data: ProjectProgressUpdate, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.SITE_ENGINEER]))):
    existing = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Project not found")
    update = {"progress_percentage": data.progress_percentage}
    if data.actual_cost is not None:
        update["actual_cost"] = data.actual_cost
    await db.projects.update_one({"id": project_id}, {"$set": update})
    updated = await db.projects.find_one({"id": project_id}, {"_id": 0})
    return updated

@api_router.get("/projects/{project_id}/summary")
async def get_project_summary(project_id: str, current_user: User = Depends(get_current_user)):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    tasks = await db.tasks.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    dprs = await db.dprs.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    billings = await db.billings.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    cvrs = await db.cvrs.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    pos = await db.purchase_orders.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    attendance = await db.attendance.find({"project_id": project_id}, {"_id": 0}).to_list(1000)

    total_tasks = len(tasks)
    completed_tasks = len([t for t in tasks if t.get('status') == 'completed'])
    in_progress_tasks = len([t for t in tasks if t.get('status') == 'in_progress'])
    total_billed = sum(b.get('total_amount', 0) for b in billings)
    total_po = sum(p.get('total', 0) for p in pos)
    total_cvr_work = sum(c.get('work_done_value', 0) for c in cvrs)
    labor_days = len([a for a in attendance if a.get('status') == 'present'])

    return {
        "project": project,
        "tasks": {"total": total_tasks, "completed": completed_tasks, "in_progress": in_progress_tasks, "pending": total_tasks - completed_tasks - in_progress_tasks},
        "dprs": {"total": len(dprs), "latest": dprs[-1] if dprs else None},
        "financial": {"total_billed": total_billed, "total_po_value": total_po, "total_cvr_work": total_cvr_work, "budget": project.get('budget', 0), "actual_cost": project.get('actual_cost', 0), "variance": project.get('budget', 0) - project.get('actual_cost', 0)},
        "workforce": {"labor_days": labor_days, "attendance_records": len(attendance)},
        "procurement": {"total_pos": len(pos), "total_po_value": total_po}
    }

# ==================== TASK ROUTES (WBS) ====================

@api_router.post("/tasks", response_model=Task)
async def create_task(task_data: TaskCreate, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.SITE_ENGINEER]))):
    task = Task(**task_data.model_dump())
    doc = task.model_dump()
    await db.tasks.insert_one(doc)
    return task

@api_router.get("/tasks", response_model=List[Task])
async def get_tasks(project_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {"project_id": project_id} if project_id else {}
    tasks = await db.tasks.find(query, {"_id": 0}).to_list(1000)
    return tasks

@api_router.put("/tasks/{task_id}", response_model=Task)
async def update_task(task_id: str, task_data: TaskCreate, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.SITE_ENGINEER]))):
    update_dict = task_data.model_dump()
    await db.tasks.update_one({"id": task_id}, {"$set": update_dict})
    updated = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    if not updated:
        raise HTTPException(status_code=404, detail="Task not found")
    return Task(**updated)

class TaskStatusUpdate(BaseModel):
    status: str
    progress: Optional[float] = None

@api_router.patch("/tasks/{task_id}/status")
async def update_task_status(task_id: str, data: TaskStatusUpdate, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.SITE_ENGINEER]))):
    existing = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Task not found")
    update = {"status": data.status}
    if data.progress is not None:
        update["progress"] = data.progress
    elif data.status == "completed":
        update["progress"] = 100.0
    await db.tasks.update_one({"id": task_id}, {"$set": update})
    updated = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    return updated

@api_router.delete("/tasks/{task_id}")
async def delete_task(task_id: str, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.SITE_ENGINEER]))):
    result = await db.tasks.delete_one({"id": task_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Task not found")
    return {"message": "Task deleted"}

# ==================== DPR ROUTES ====================

@api_router.post("/dpr", response_model=DPR)
async def create_dpr(dpr_data: DPRCreate, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.SITE_ENGINEER]))):
    dpr = DPR(**dpr_data.model_dump(), created_by=current_user.id)
    doc = dpr.model_dump()
    await db.dprs.insert_one(doc)
    return dpr

@api_router.get("/dpr", response_model=List[DPR])
async def get_dprs(project_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {"project_id": project_id} if project_id else {}
    dprs = await db.dprs.find(query, {"_id": 0}).to_list(1000)
    return dprs

# ==================== CVR ROUTES ====================

@api_router.post("/cvr", response_model=CVR)
async def create_cvr(cvr_data: CVRCreate, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.FINANCE]))):
    cvr = CVR(**cvr_data.model_dump())
    cvr.variance = cvr.contracted_value - cvr.work_done_value
    doc = cvr.model_dump()
    await db.cvrs.insert_one(doc)
    return cvr

@api_router.get("/cvr", response_model=List[CVR])
async def get_cvrs(project_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {"project_id": project_id} if project_id else {}
    cvrs = await db.cvrs.find(query, {"_id": 0}).to_list(1000)
    return cvrs

# ==================== BILLING ROUTES ====================

@api_router.post("/billing", response_model=Billing)
async def create_billing(billing_data: BillingCreate, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.FINANCE]))):
    billing = Billing(**billing_data.model_dump())
    billing.gst_amount = billing.amount * billing.gst_rate / 100
    billing.total_amount = billing.amount + billing.gst_amount
    doc = billing.model_dump()
    await db.billings.insert_one(doc)
    return billing

@api_router.get("/billing", response_model=List[Billing])
async def get_billings(project_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {"project_id": project_id} if project_id else {}
    billings = await db.billings.find(query, {"_id": 0}).to_list(1000)
    return billings

@api_router.put("/billing/{billing_id}/status")
async def update_billing_status(billing_id: str, status: str, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.FINANCE]))):
    await db.billings.update_one({"id": billing_id}, {"$set": {"status": status}})
    return {"message": "Status updated"}

@api_router.get("/billing/{billing_id}")
async def get_billing(billing_id: str, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.FINANCE]))):
    bill = await db.billings.find_one({"id": billing_id}, {"_id": 0})
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")
    return bill

@api_router.delete("/billing/{billing_id}")
async def delete_billing(billing_id: str, current_user: User = Depends(check_role([UserRole.ADMIN]))):
    result = await db.billings.delete_one({"id": billing_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Bill not found")
    return {"message": "Bill deleted"}

class BillingStatusUpdate(BaseModel):
    status: str

@api_router.patch("/billing/{billing_id}/status")
async def patch_billing_status(billing_id: str, data: BillingStatusUpdate, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.FINANCE]))):
    existing = await db.billings.find_one({"id": billing_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Bill not found")
    await db.billings.update_one({"id": billing_id}, {"$set": {"status": data.status}})
    updated = await db.billings.find_one({"id": billing_id}, {"_id": 0})
    return updated

@api_router.delete("/cvr/{cvr_id}")
async def delete_cvr(cvr_id: str, current_user: User = Depends(check_role([UserRole.ADMIN]))):
    result = await db.cvrs.delete_one({"id": cvr_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="CVR not found")
    return {"message": "CVR deleted"}

@api_router.get("/financial/dashboard")
async def get_financial_dashboard(current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.FINANCE]))):
    billings = await db.billings.find({}, {"_id": 0}).to_list(1000)
    cvrs = await db.cvrs.find({}, {"_id": 0}).to_list(1000)
    projects = await db.projects.find({}, {"_id": 0}).to_list(1000)

    total_billed = sum(b.get('total_amount', 0) for b in billings)
    total_gst = sum(b.get('gst_amount', 0) for b in billings)
    pending_amount = sum(b.get('total_amount', 0) for b in billings if b.get('status') == 'pending')
    approved_amount = sum(b.get('total_amount', 0) for b in billings if b.get('status') == 'approved')
    paid_amount = sum(b.get('total_amount', 0) for b in billings if b.get('status') == 'paid')
    total_received = sum(c.get('received_value', 0) for c in cvrs)
    total_retention = sum(c.get('retention_held', 0) for c in cvrs)
    total_contracted = sum(c.get('contracted_value', 0) for c in cvrs)
    total_work_done = sum(c.get('work_done_value', 0) for c in cvrs)
    total_budget = sum(p.get('budget', 0) for p in projects)
    total_spent = sum(p.get('actual_cost', 0) for p in projects)

    collection_eff = round((total_received / total_billed * 100) if total_billed > 0 else 0, 1)
    cpi = round((total_work_done / total_spent) if total_spent > 0 else 0, 2)

    # Project-wise financial breakdown
    project_breakdown = []
    for p in projects:
        pid = p.get('id')
        p_bills = [b for b in billings if b.get('project_id') == pid]
        p_cvrs = [c for c in cvrs if c.get('project_id') == pid]
        project_breakdown.append({
            "project_id": pid,
            "project_name": p.get('name'),
            "budget": p.get('budget', 0),
            "actual_cost": p.get('actual_cost', 0),
            "total_billed": sum(b.get('total_amount', 0) for b in p_bills),
            "bills_count": len(p_bills),
            "received": sum(c.get('received_value', 0) for c in p_cvrs),
            "variance": p.get('budget', 0) - p.get('actual_cost', 0)
        })

    # Bills by status
    bills_by_status = {"pending": 0, "approved": 0, "paid": 0}
    for b in billings:
        s = b.get('status', 'pending')
        bills_by_status[s] = bills_by_status.get(s, 0) + 1

    # Bills by type
    bills_by_type = {"running": 0, "final": 0, "advance": 0}
    for b in billings:
        t = b.get('bill_type', 'running')
        bills_by_type[t] = bills_by_type.get(t, 0) + 1

    return {
        "summary": {
            "total_billed": total_billed,
            "total_gst": total_gst,
            "pending_collection": pending_amount,
            "approved_amount": approved_amount,
            "paid_amount": paid_amount,
            "total_received": total_received,
            "total_retention": total_retention,
            "collection_efficiency": collection_eff,
            "total_budget": total_budget,
            "total_spent": total_spent,
            "total_contracted": total_contracted,
            "total_work_done": total_work_done,
            "cpi": cpi,
            "total_bills": len(billings),
            "total_cvrs": len(cvrs)
        },
        "bills_by_status": bills_by_status,
        "bills_by_type": bills_by_type,
        "project_breakdown": sorted(project_breakdown, key=lambda x: x['total_billed'], reverse=True)
    }

# ==================== VENDOR ROUTES ====================

@api_router.post("/vendors", response_model=Vendor)
async def create_vendor(vendor_data: VendorCreate, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.PROCUREMENT]))):
    vendor = Vendor(**vendor_data.model_dump())
    doc = vendor.model_dump()
    await db.vendors.insert_one(doc)
    return vendor

@api_router.get("/vendors", response_model=List[Vendor])
async def get_vendors(category: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {"is_active": True}
    if category:
        query["category"] = category
    vendors = await db.vendors.find(query, {"_id": 0}).to_list(1000)
    return vendors

@api_router.get("/vendors/{vendor_id}", response_model=Vendor)
async def get_vendor(vendor_id: str, current_user: User = Depends(get_current_user)):
    vendor = await db.vendors.find_one({"id": vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return Vendor(**vendor)

@api_router.get("/vendors/{vendor_id}/detail")
async def get_vendor_detail(vendor_id: str, current_user: User = Depends(get_current_user)):
    vendor = await db.vendors.find_one({"id": vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    pos = await db.purchase_orders.find({"vendor_id": vendor_id}, {"_id": 0}).to_list(1000)
    grns = await db.grns.find({}, {"_id": 0}).to_list(1000)
    po_ids = {po.get("id") for po in pos}
    vendor_grns = [g for g in grns if g.get("po_id") in po_ids]
    total_po_value = sum(po.get("total", 0) for po in pos)
    po_by_status = {}
    for po in pos:
        s = po.get("status", "pending")
        po_by_status[s] = po_by_status.get(s, 0) + 1
    return {
        "vendor": vendor,
        "purchase_orders": pos,
        "grns": vendor_grns,
        "stats": {
            "total_pos": len(pos),
            "total_po_value": total_po_value,
            "total_grns": len(vendor_grns),
            "po_by_status": po_by_status,
            "avg_po_value": round(total_po_value / len(pos)) if pos else 0
        }
    }

@api_router.put("/vendors/{vendor_id}", response_model=Vendor)
async def update_vendor(vendor_id: str, vendor_data: VendorCreate, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.PROCUREMENT]))):
    update_dict = vendor_data.model_dump()
    await db.vendors.update_one({"id": vendor_id}, {"$set": update_dict})
    updated = await db.vendors.find_one({"id": vendor_id}, {"_id": 0})
    if not updated:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return Vendor(**updated)

class VendorRating(BaseModel):
    rating: float

@api_router.patch("/vendors/{vendor_id}/rating")
async def rate_vendor(vendor_id: str, data: VendorRating, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.PROCUREMENT]))):
    existing = await db.vendors.find_one({"id": vendor_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Vendor not found")
    await db.vendors.update_one({"id": vendor_id}, {"$set": {"rating": data.rating}})
    updated = await db.vendors.find_one({"id": vendor_id}, {"_id": 0})
    return updated

@api_router.patch("/vendors/{vendor_id}/deactivate")
async def deactivate_vendor(vendor_id: str, current_user: User = Depends(check_role([UserRole.ADMIN]))):
    existing = await db.vendors.find_one({"id": vendor_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Vendor not found")
    await db.vendors.update_one({"id": vendor_id}, {"$set": {"is_active": False}})
    return {"message": "Vendor deactivated"}

# ==================== PURCHASE ORDER ROUTES ====================

@api_router.post("/purchase-orders", response_model=PurchaseOrder)
async def create_purchase_order(po_data: PurchaseOrderCreate, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.PROCUREMENT]))):
    count = await db.purchase_orders.count_documents({})
    po_number = f"PO-{datetime.now().strftime('%Y%m')}-{count + 1:04d}"
    items = [item.model_dump() for item in po_data.items]
    subtotal = sum(item['quantity'] * item['rate'] for item in items)
    gst_amount = subtotal * 0.18
    po = PurchaseOrder(
        po_number=po_number, project_id=po_data.project_id, vendor_id=po_data.vendor_id,
        po_date=po_data.po_date, delivery_date=po_data.delivery_date, items=items,
        terms=po_data.terms, subtotal=subtotal, gst_amount=gst_amount, total=subtotal + gst_amount
    )
    doc = po.model_dump()
    await db.purchase_orders.insert_one(doc)
    return po

@api_router.get("/purchase-orders")
async def get_purchase_orders(project_id: Optional[str] = None, vendor_id: Optional[str] = None, status: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if project_id: query["project_id"] = project_id
    if vendor_id: query["vendor_id"] = vendor_id
    if status: query["status"] = status
    pos = await db.purchase_orders.find(query, {"_id": 0}).to_list(1000)
    return pos

@api_router.get("/purchase-orders/{po_id}")
async def get_purchase_order(po_id: str, current_user: User = Depends(get_current_user)):
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="PO not found")
    vendor = await db.vendors.find_one({"id": po.get("vendor_id")}, {"_id": 0})
    project = await db.projects.find_one({"id": po.get("project_id")}, {"_id": 0})
    grns = await db.grns.find({"po_id": po_id}, {"_id": 0}).to_list(100)
    # 3-way matching
    total_ordered = {i: item.get("quantity", 0) for i, item in enumerate(po.get("items", []))}
    total_received = {}
    for grn in grns:
        for gi in grn.get("items", []):
            idx = gi.get("po_item_index", 0)
            total_received[idx] = total_received.get(idx, 0) + gi.get("received_quantity", 0)
    matching = []
    for i, item in enumerate(po.get("items", [])):
        ordered = item.get("quantity", 0)
        received = total_received.get(i, 0)
        matching.append({
            "item_index": i,
            "description": item.get("description"),
            "ordered": ordered,
            "received": received,
            "pending": ordered - received,
            "status": "complete" if received >= ordered else "partial" if received > 0 else "pending"
        })
    return {
        "po": po,
        "vendor": vendor,
        "project": project,
        "grns": grns,
        "matching": matching
    }

class POStatusUpdate(BaseModel):
    status: str

@api_router.patch("/purchase-orders/{po_id}/status")
async def patch_po_status(po_id: str, data: POStatusUpdate, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.PROCUREMENT]))):
    existing = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="PO not found")
    await db.purchase_orders.update_one({"id": po_id}, {"$set": {"status": data.status}})
    updated = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    return updated

@api_router.delete("/purchase-orders/{po_id}")
async def delete_po(po_id: str, current_user: User = Depends(check_role([UserRole.ADMIN]))):
    result = await db.purchase_orders.delete_one({"id": po_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="PO not found")
    return {"message": "PO deleted"}

@api_router.get("/procurement/dashboard")
async def get_procurement_dashboard(current_user: User = Depends(get_current_user)):
    vendors = await db.vendors.find({"is_active": True}, {"_id": 0}).to_list(1000)
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(1000)
    grns = await db.grns.find({}, {"_id": 0}).to_list(1000)
    total_po_value = sum(po.get("total", 0) for po in pos)
    pending_pos = len([p for p in pos if p.get("status") == "pending"])
    approved_pos = len([p for p in pos if p.get("status") == "approved"])
    by_category = {}
    for v in vendors:
        c = v.get("category", "other")
        by_category[c] = by_category.get(c, 0) + 1
    # Top vendor by PO value
    vendor_po_map = {}
    for po in pos:
        vid = po.get("vendor_id")
        vendor_po_map[vid] = vendor_po_map.get(vid, 0) + po.get("total", 0)
    top_vendor_id = max(vendor_po_map, key=vendor_po_map.get) if vendor_po_map else None
    top_vendor = next((v for v in vendors if v.get("id") == top_vendor_id), None) if top_vendor_id else None
    return {
        "vendors": {"total": len(vendors), "by_category": by_category},
        "purchase_orders": {"total": len(pos), "total_value": total_po_value, "pending": pending_pos, "approved": approved_pos, "delivered": len([p for p in pos if p.get("status") == "delivered"]), "closed": len([p for p in pos if p.get("status") == "closed"])},
        "grns": {"total": len(grns)},
        "top_vendor": {"name": top_vendor.get("name") if top_vendor else "-", "value": vendor_po_map.get(top_vendor_id, 0) if top_vendor_id else 0}
    }

# ==================== GRN ROUTES ====================

@api_router.post("/grn", response_model=GRN)
async def create_grn(grn_data: GRNCreate, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.PROCUREMENT, UserRole.SITE_ENGINEER]))):
    count = await db.grns.count_documents({})
    grn_number = f"GRN-{datetime.now().strftime('%Y%m')}-{count + 1:04d}"
    items = [item.model_dump() for item in grn_data.items]
    grn = GRN(grn_number=grn_number, po_id=grn_data.po_id, grn_date=grn_data.grn_date, items=items, notes=grn_data.notes)
    doc = grn.model_dump()
    await db.grns.insert_one(doc)
    return grn

@api_router.get("/grn")
async def get_grns(po_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {"po_id": po_id} if po_id else {}
    grns = await db.grns.find(query, {"_id": 0}).to_list(1000)
    return grns

@api_router.delete("/grn/{grn_id}")
async def delete_grn(grn_id: str, current_user: User = Depends(check_role([UserRole.ADMIN]))):
    result = await db.grns.delete_one({"id": grn_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="GRN not found")
    return {"message": "GRN deleted"}

# ==================== EMPLOYEE ROUTES ====================

@api_router.post("/employees", response_model=Employee)
async def create_employee(employee_data: EmployeeCreate, current_user: User = Depends(check_role([UserRole.ADMIN]))):
    emp_dict = employee_data.model_dump()
    user_id = emp_dict.pop('user_id', None)
    create_user_account = emp_dict.pop('create_user_account', False)
    
    # If create_user_account is True, create a user account with default password
    if create_user_account and not user_id:
        # Check if user with this email already exists
        existing_user = await db.users.find_one({"email": employee_data.email})
        if existing_user:
            user_id = existing_user.get("id")
        else:
            # Create a new user with default password (employee should change it)
            default_password = "Welcome@123"
            new_user = User(
                email=employee_data.email,
                name=employee_data.name,
                role=UserRole.SITE_ENGINEER,  # Default role
                phone=employee_data.phone,
                department=employee_data.department
            )
            user_doc = {**new_user.model_dump(), "password": get_password_hash(default_password)}
            await db.users.insert_one(user_doc)
            user_id = new_user.id
    
    # Create employee with optional user_id link
    emp_dict['user_id'] = user_id
    employee = Employee(**emp_dict)
    doc = employee.model_dump()
    await db.employees.insert_one(doc)
    return employee

@api_router.get("/employees")
async def get_employees(department: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {"is_active": True}
    if department:
        query["department"] = department
    employees = await db.employees.find(query, {"_id": 0}).to_list(1000)
    return employees

@api_router.get("/employees/{employee_id}")
async def get_employee(employee_id: str, current_user: User = Depends(get_current_user)):
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    return employee

@api_router.get("/employees/{employee_id}/detail")
async def get_employee_detail(employee_id: str, current_user: User = Depends(get_current_user)):
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    att = await db.attendance.find({"employee_id": employee_id}, {"_id": 0}).sort("date", -1).to_list(1000)
    pays = await db.payrolls.find({"employee_id": employee_id}, {"_id": 0}).sort("month", -1).to_list(1000)
    present = len([a for a in att if a.get("status") == "present"])
    absent = len([a for a in att if a.get("status") == "absent"])
    half = len([a for a in att if a.get("status") == "half_day"])
    leave = len([a for a in att if a.get("status") == "leave"])
    total_ot = sum(a.get("overtime_hours", 0) for a in att)
    total_paid = sum(p.get("net_salary", 0) for p in pays)
    att_rate = round((present / len(att) * 100) if att else 0, 1)
    return {
        "employee": employee,
        "attendance": att[:30],
        "payrolls": pays,
        "stats": {
            "total_attendance": len(att), "present": present, "absent": absent,
            "half_day": half, "leave": leave, "attendance_rate": att_rate,
            "total_overtime": total_ot, "total_payrolls": len(pays), "total_paid": total_paid
        }
    }

@api_router.put("/employees/{employee_id}")
async def update_employee(employee_id: str, employee_data: EmployeeCreate, current_user: User = Depends(check_role([UserRole.ADMIN]))):
    update_dict = employee_data.model_dump()
    await db.employees.update_one({"id": employee_id}, {"$set": update_dict})
    updated = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not updated:
        raise HTTPException(status_code=404, detail="Employee not found")
    return updated

@api_router.patch("/employees/{employee_id}/deactivate")
async def deactivate_employee(employee_id: str, current_user: User = Depends(check_role([UserRole.ADMIN]))):
    existing = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Employee not found")
    await db.employees.update_one({"id": employee_id}, {"$set": {"is_active": False}})
    return {"message": "Employee deactivated"}

class LinkEmployeeUser(BaseModel):
    employee_id: str
    user_id: Optional[str] = None  # If None, will create new user

@api_router.post("/employees/{employee_id}/link-user")
async def link_employee_to_user(
    employee_id: str, 
    link_data: LinkEmployeeUser,
    current_user: User = Depends(check_role([UserRole.ADMIN]))
):
    """Link an employee to an existing user account or create a new one"""
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    user_id = link_data.user_id
    
    if not user_id:
        # Create a new user account for this employee
        existing_user = await db.users.find_one({"email": employee["email"]})
        if existing_user:
            user_id = existing_user.get("id")
        else:
            default_password = "Welcome@123"
            new_user = User(
                email=employee["email"],
                name=employee["name"],
                role=UserRole.SITE_ENGINEER,
                phone=employee.get("phone"),
                department=employee.get("department")
            )
            user_doc = {**new_user.model_dump(), "password": get_password_hash(default_password)}
            await db.users.insert_one(user_doc)
            user_id = new_user.id
    else:
        # Verify user exists
        user = await db.users.find_one({"id": user_id})
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
    
    await db.employees.update_one({"id": employee_id}, {"$set": {"user_id": user_id}})
    return {"message": "Employee linked to user account", "user_id": user_id}

@api_router.delete("/employees/{employee_id}/unlink-user")
async def unlink_employee_from_user(employee_id: str, current_user: User = Depends(check_role([UserRole.ADMIN]))):
    """Remove the link between an employee and their user account"""
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    await db.employees.update_one({"id": employee_id}, {"$unset": {"user_id": ""}})
    return {"message": "Employee unlinked from user account"}

@api_router.get("/rbac/employees")
async def list_employees_with_user_info(current_user: User = Depends(require_admin())):
    """Get all employees with their linked user and role information for RBAC management"""
    employees = await db.employees.find({"is_active": True}, {"_id": 0}).to_list(1000)
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    roles = await db.roles.find({}, {"_id": 0}).to_list(1000)
    
    users_map = {u["id"]: u for u in users}
    roles_map = {r["id"]: r["name"] for r in roles}
    
    result = []
    for emp in employees:
        user_data = None
        if emp.get("user_id"):
            user = users_map.get(emp["user_id"])
            if user:
                user_data = {
                    "id": user.get("id"),
                    "email": user.get("email"),
                    "role": user.get("role"),
                    "role_id": user.get("role_id"),
                    "role_name": roles_map.get(user.get("role_id"))
                }
        
        result.append({
            "id": emp.get("id"),
            "name": emp.get("name"),
            "employee_code": emp.get("employee_code"),
            "email": emp.get("email"),
            "department": emp.get("department"),
            "designation": emp.get("designation"),
            "user_id": emp.get("user_id"),
            "user_data": user_data,
            "has_user_account": emp.get("user_id") is not None
        })
    
    return result

# ==================== ATTENDANCE ROUTES ====================

@api_router.post("/attendance")
async def create_attendance(attendance_data: AttendanceCreate, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.SITE_ENGINEER]))):
    attendance = Attendance(**attendance_data.model_dump())
    doc = attendance.model_dump()
    await db.attendance.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.get("/attendance")
async def get_attendance(employee_id: Optional[str] = None, project_id: Optional[str] = None, date: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if employee_id: query["employee_id"] = employee_id
    if project_id: query["project_id"] = project_id
    if date: query["date"] = date
    attendance = await db.attendance.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return attendance

@api_router.delete("/attendance/{att_id}")
async def delete_attendance(att_id: str, current_user: User = Depends(check_role([UserRole.ADMIN]))):
    result = await db.attendance.delete_one({"id": att_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Attendance not found")
    return {"message": "Deleted"}

# ==================== PAYROLL ROUTES ====================

@api_router.post("/payroll")
async def create_payroll(payroll_data: PayrollCreate, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.FINANCE]))):
    payroll = Payroll(**payroll_data.model_dump())
    payroll.gross_salary = payroll.basic_salary + payroll.hra + payroll.overtime_pay + payroll.other_allowances
    payroll.total_deductions = payroll.pf_deduction + payroll.esi_deduction + payroll.tds + payroll.other_deductions
    payroll.net_salary = payroll.gross_salary - payroll.total_deductions
    doc = payroll.model_dump()
    await db.payrolls.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.get("/payroll")
async def get_payrolls(employee_id: Optional[str] = None, month: Optional[str] = None, status: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if employee_id: query["employee_id"] = employee_id
    if month: query["month"] = month
    if status: query["status"] = status
    payrolls = await db.payrolls.find(query, {"_id": 0}).to_list(1000)
    return payrolls

class PayrollStatusUpdate(BaseModel):
    status: str

@api_router.patch("/payroll/{payroll_id}/status")
async def update_payroll_status(payroll_id: str, data: PayrollStatusUpdate, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.FINANCE]))):
    existing = await db.payrolls.find_one({"id": payroll_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Payroll not found")
    await db.payrolls.update_one({"id": payroll_id}, {"$set": {"status": data.status}})
    updated = await db.payrolls.find_one({"id": payroll_id}, {"_id": 0})
    return updated

@api_router.delete("/payroll/{payroll_id}")
async def delete_payroll(payroll_id: str, current_user: User = Depends(check_role([UserRole.ADMIN]))):
    result = await db.payrolls.delete_one({"id": payroll_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Payroll not found")
    return {"message": "Deleted"}

@api_router.get("/hrms/dashboard")
async def get_hrms_dashboard(current_user: User = Depends(get_current_user)):
    employees = await db.employees.find({"is_active": True}, {"_id": 0}).to_list(1000)
    attendance = await db.attendance.find({}, {"_id": 0}).to_list(5000)
    payrolls = await db.payrolls.find({}, {"_id": 0}).to_list(1000)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    today_att = [a for a in attendance if a.get("date") == today]
    present_today = len([a for a in today_att if a.get("status") == "present"])
    by_dept = {}
    for e in employees:
        d = e.get("department", "Other")
        by_dept[d] = by_dept.get(d, 0) + 1
    total_salary = sum(e.get("basic_salary", 0) + e.get("hra", 0) for e in employees)
    total_payroll = sum(p.get("net_salary", 0) for p in payrolls)
    pending_payrolls = len([p for p in payrolls if p.get("status") == "pending"])
    total_ot = sum(a.get("overtime_hours", 0) for a in attendance)
    all_present = len([a for a in attendance if a.get("status") == "present"])
    att_rate = round((all_present / len(attendance) * 100) if attendance else 0, 1)
    return {
        "employees": {"total": len(employees), "by_department": by_dept, "monthly_salary_budget": total_salary},
        "attendance": {"present_today": present_today, "total_today": len(today_att), "overall_rate": att_rate, "total_overtime": total_ot},
        "payroll": {"total_disbursed": total_payroll, "pending": pending_payrolls, "total_processed": len(payrolls)}
    }

# ==================== GST ROUTES ====================

@api_router.post("/gst-returns", response_model=GSTReturn)
async def create_gst_return(gst_data: GSTReturnCreate, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.FINANCE]))):
    gst_return = GSTReturn(**gst_data.model_dump())
    gst_return.tax_payable = gst_return.cgst + gst_return.sgst + gst_return.igst - gst_return.itc_claimed
    doc = gst_return.model_dump()
    await db.gst_returns.insert_one(doc)
    return gst_return

@api_router.get("/gst-returns", response_model=List[GSTReturn])
async def get_gst_returns(current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.FINANCE]))):
    gst_returns = await db.gst_returns.find({}, {"_id": 0}).to_list(1000)
    return gst_returns

# ==================== RERA ROUTES ====================

@api_router.post("/rera-projects", response_model=RERAProject)
async def create_rera_project(rera_data: RERAProjectCreate, current_user: User = Depends(check_role([UserRole.ADMIN]))):
    rera_project = RERAProject(**rera_data.model_dump())
    doc = rera_project.model_dump()
    await db.rera_projects.insert_one(doc)
    return rera_project

@api_router.get("/rera-projects", response_model=List[RERAProject])
async def get_rera_projects(current_user: User = Depends(get_current_user)):
    rera_projects = await db.rera_projects.find({}, {"_id": 0}).to_list(1000)
    return rera_projects

# ==================== GST SETTINGS API ====================

class GSTCredentialsCreate(BaseModel):
    gstin: str
    username: str
    password: str
    client_id: str
    client_secret: str
    nic_url: str = "https://einv-apisandbox.nic.in"
    is_sandbox: bool = True

class GSTCredentialsResponse(BaseModel):
    gstin: str
    username: str
    client_id: str
    nic_url: str
    is_sandbox: bool
    is_configured: bool = True
    last_updated: Optional[str] = None

def encrypt_value(value: str) -> str:
    return fernet.encrypt(value.encode()).decode()

def decrypt_value(value: str) -> str:
    return fernet.decrypt(value.encode()).decode()

@api_router.post("/settings/gst-credentials")
async def save_gst_credentials(creds: GSTCredentialsCreate, current_user: User = Depends(check_role([UserRole.ADMIN]))):
    existing = await db.gst_settings.find_one({}, {"_id": 0})
    
    password_enc = encrypt_value(creds.password) if creds.password != "___unchanged___" else (existing or {}).get("password_enc", "")
    secret_enc = encrypt_value(creds.client_secret) if creds.client_secret != "___unchanged___" else (existing or {}).get("client_secret_enc", "")
    
    doc = {
        "gstin": creds.gstin,
        "username": creds.username,
        "password_enc": password_enc,
        "client_id": creds.client_id,
        "client_secret_enc": secret_enc,
        "nic_url": creds.nic_url,
        "is_sandbox": creds.is_sandbox,
        "updated_by": current_user.id,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.gst_settings.update_one({}, {"$set": doc}, upsert=True)
    return {"message": "GST credentials saved", "gstin": creds.gstin}

@api_router.get("/settings/gst-credentials")
async def get_gst_credentials(current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.FINANCE]))):
    settings = await db.gst_settings.find_one({}, {"_id": 0})
    if not settings:
        return {"is_configured": False}
    return GSTCredentialsResponse(
        gstin=settings["gstin"],
        username=settings["username"],
        client_id=settings["client_id"],
        nic_url=settings["nic_url"],
        is_sandbox=settings.get("is_sandbox", True),
        is_configured=True,
        last_updated=settings.get("updated_at")
    ).model_dump()

@api_router.delete("/settings/gst-credentials")
async def delete_gst_credentials(current_user: User = Depends(check_role([UserRole.ADMIN]))):
    await db.gst_settings.delete_many({})
    return {"message": "GST credentials deleted"}

@api_router.post("/settings/gst-credentials/test")
async def test_gst_connection(current_user: User = Depends(check_role([UserRole.ADMIN]))):
    settings = await db.gst_settings.find_one({}, {"_id": 0})
    if not settings:
        raise HTTPException(status_code=400, detail="GST credentials not configured")
    try:
        nic_url = settings["nic_url"]
        async with httpx.AsyncClient(timeout=15.0) as client_http:
            auth_response = await client_http.post(
                f"{nic_url}/eivital/v1.04/auth",
                json={
                    "UserName": settings["username"],
                    "Password": decrypt_value(settings["password_enc"]),
                    "AppKey": settings["client_id"],
                    "ForceRefreshAccessToken": "true"
                },
                headers={
                    "client_id": settings["client_id"],
                    "client_secret": decrypt_value(settings["client_secret_enc"]),
                    "gstin": settings["gstin"]
                }
            )
            if auth_response.status_code == 200:
                data = auth_response.json()
                if data.get("Status") == 1:
                    return {"status": "connected", "message": "NIC Portal connection successful"}
                return {"status": "auth_failed", "message": data.get("ErrorDetails", [{}])[0].get("ErrorMessage", "Authentication failed")}
            return {"status": "error", "message": f"NIC Portal returned status {auth_response.status_code}"}
    except httpx.ConnectError:
        return {"status": "unreachable", "message": "Cannot reach NIC portal. Check URL and network."}
    except Exception as e:
        return {"status": "error", "message": str(e)}

# ==================== E-INVOICE API ====================

async def get_nic_auth_token():
    """Get auth token from NIC portal using stored credentials"""
    settings = await db.gst_settings.find_one({}, {"_id": 0})
    if not settings:
        return None, "GST credentials not configured. Go to Settings > GST Integration to set up."
    try:
        nic_url = settings["nic_url"]
        async with httpx.AsyncClient(timeout=15.0) as client_http:
            auth_response = await client_http.post(
                f"{nic_url}/eivital/v1.04/auth",
                json={
                    "UserName": settings["username"],
                    "Password": decrypt_value(settings["password_enc"]),
                    "AppKey": settings["client_id"],
                    "ForceRefreshAccessToken": "true"
                },
                headers={
                    "client_id": settings["client_id"],
                    "client_secret": decrypt_value(settings["client_secret_enc"]),
                    "gstin": settings["gstin"]
                }
            )
            if auth_response.status_code == 200:
                data = auth_response.json()
                if data.get("Status") == 1:
                    return {
                        "token": data["Data"]["AuthToken"],
                        "sek": data["Data"]["Sek"],
                        "gstin": settings["gstin"],
                        "nic_url": nic_url
                    }, None
                error_msg = data.get("ErrorDetails", [{}])[0].get("ErrorMessage", "Auth failed")
                return None, error_msg
            return None, f"NIC auth returned {auth_response.status_code}"
    except Exception as e:
        return None, str(e)

def build_nic_invoice_payload(invoice_data: EInvoiceCreate) -> dict:
    """Build FORM INV-01 standard payload for NIC portal"""
    items_list = []
    for item in invoice_data.items:
        items_list.append({
            "SlNo": str(item.sl_no),
            "PrdDesc": item.item_description,
            "IsServc": "N",
            "HsnCd": item.hsn_code,
            "Qty": item.quantity,
            "Unit": item.unit,
            "UnitPrice": item.unit_price,
            "Discount": item.discount,
            "TotAmt": item.quantity * item.unit_price,
            "AssAmt": item.taxable_value,
            "GstRt": item.gst_rate,
            "CgstAmt": item.cgst_amount,
            "SgstAmt": item.sgst_amount,
            "IgstAmt": item.igst_amount,
            "CesAmt": item.cess_amount,
            "TotItemVal": item.total_item_value
        })
    
    payload = {
        "Version": "1.1",
        "TranDtls": {
            "TaxSch": "GST",
            "SupTyp": invoice_data.supply_type,
            "RegRev": "N",
            "IgstOnIntra": "N"
        },
        "DocDtls": {
            "Typ": invoice_data.document_type,
            "No": invoice_data.document_number,
            "Dt": invoice_data.document_date
        },
        "SellerDtls": {
            "Gstin": invoice_data.seller_gstin,
            "LglNm": invoice_data.seller_legal_name,
            "TrdNm": invoice_data.seller_trade_name or invoice_data.seller_legal_name,
            "Addr1": invoice_data.seller_address,
            "Loc": invoice_data.seller_location,
            "Pin": int(invoice_data.seller_pincode),
            "Stcd": invoice_data.seller_state_code
        },
        "BuyerDtls": {
            "Gstin": invoice_data.buyer_gstin,
            "LglNm": invoice_data.buyer_legal_name,
            "TrdNm": invoice_data.buyer_trade_name or invoice_data.buyer_legal_name,
            "Pos": invoice_data.buyer_pos,
            "Addr1": invoice_data.buyer_address,
            "Loc": invoice_data.buyer_location,
            "Pin": int(invoice_data.buyer_pincode),
            "Stcd": invoice_data.buyer_state_code
        },
        "ItemList": items_list,
        "ValDtls": {
            "AssVal": invoice_data.total_taxable_value,
            "CgstVal": invoice_data.total_cgst,
            "SgstVal": invoice_data.total_sgst,
            "IgstVal": invoice_data.total_igst,
            "CesVal": invoice_data.total_cess,
            "Discount": invoice_data.total_discount,
            "OthChrg": invoice_data.other_charges,
            "RndOffAmt": invoice_data.round_off,
            "TotInvVal": invoice_data.total_invoice_value
        },
        "PayDtls": {
            "Nm": invoice_data.buyer_legal_name,
            "Mode": invoice_data.payment_mode
        }
    }
    
    if invoice_data.dispatch_from_name:
        payload["DispDtls"] = {
            "Nm": invoice_data.dispatch_from_name,
            "Addr1": invoice_data.dispatch_from_address,
            "Loc": invoice_data.dispatch_from_location,
            "Pin": int(invoice_data.dispatch_from_pincode) if invoice_data.dispatch_from_pincode else 0,
            "Stcd": invoice_data.dispatch_from_state_code
        }
    
    if invoice_data.ship_to_gstin:
        payload["ShipDtls"] = {
            "Gstin": invoice_data.ship_to_gstin,
            "LglNm": invoice_data.ship_to_legal_name,
            "Addr1": invoice_data.ship_to_address,
            "Loc": invoice_data.ship_to_location,
            "Pin": int(invoice_data.ship_to_pincode) if invoice_data.ship_to_pincode else 0,
            "Stcd": invoice_data.ship_to_state_code
        }
    
    if invoice_data.transporter_id:
        payload["EwbDtls"] = {
            "TransId": invoice_data.transporter_id,
            "TransName": invoice_data.transporter_name,
            "TransMode": invoice_data.transport_mode,
            "Distance": invoice_data.transport_distance or 0,
            "VehNo": invoice_data.vehicle_number,
            "VehType": invoice_data.vehicle_type
        }
    
    return payload

def generate_qr_base64(data: str) -> str:
    """Generate QR code as base64 string"""
    qr = qrcode.QRCode(version=1, box_size=6, border=2)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    return base64.b64encode(buffer.getvalue()).decode()

@api_router.post("/einvoice/generate")
async def generate_einvoice(invoice_data: EInvoiceCreate, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.FINANCE]))):
    """Generate E-Invoice via NIC portal or in test mode"""
    # Create the e-invoice record
    einvoice = EInvoice(
        billing_id=invoice_data.billing_id,
        document_number=invoice_data.document_number,
        document_date=invoice_data.document_date,
        document_type=invoice_data.document_type,
        supply_type=invoice_data.supply_type,
        seller_gstin=invoice_data.seller_gstin,
        seller_legal_name=invoice_data.seller_legal_name,
        buyer_gstin=invoice_data.buyer_gstin,
        buyer_legal_name=invoice_data.buyer_legal_name,
        total_taxable_value=invoice_data.total_taxable_value,
        total_cgst=invoice_data.total_cgst,
        total_sgst=invoice_data.total_sgst,
        total_igst=invoice_data.total_igst,
        total_invoice_value=invoice_data.total_invoice_value,
        items=[item.model_dump() for item in invoice_data.items],
        status="draft"
    )
    
    # Check if GST credentials are configured
    settings = await db.gst_settings.find_one({}, {"_id": 0})
    
    if settings:
        # Attempt real NIC API call
        auth_result, auth_error = await get_nic_auth_token()
        if auth_error:
            # Save as draft with error
            einvoice.status = "auth_failed"
            einvoice.error_details = f"NIC Auth Failed: {auth_error}"
            doc = einvoice.model_dump()
            await db.e_invoices.insert_one(doc)
            doc.pop("_id", None)
            return doc
        
        try:
            nic_payload = build_nic_invoice_payload(invoice_data)
            nic_url = auth_result["nic_url"]
            
            async with httpx.AsyncClient(timeout=30.0) as client_http:
                irn_response = await client_http.post(
                    f"{nic_url}/eicore/v1.03/Invoice",
                    json=nic_payload,
                    headers={
                        "client_id": settings["client_id"],
                        "client_secret": decrypt_value(settings["client_secret_enc"]),
                        "gstin": settings["gstin"],
                        "user_name": settings["username"],
                        "AuthToken": auth_result["token"],
                        "Sek": auth_result["sek"]
                    }
                )
                
                nic_data = irn_response.json()
                einvoice.nic_response = nic_data
                
                if nic_data.get("Status") == 1:
                    result_data = nic_data.get("Data", {})
                    einvoice.irn = result_data.get("Irn")
                    einvoice.ack_number = str(result_data.get("AckNo", ""))
                    einvoice.ack_date = result_data.get("AckDt")
                    einvoice.signed_invoice = result_data.get("SignedInvoice")
                    einvoice.signed_qr_code = result_data.get("SignedQRCode")
                    if einvoice.signed_qr_code:
                        einvoice.qr_code_image = generate_qr_base64(einvoice.signed_qr_code)
                    einvoice.status = "irn_generated"
                else:
                    errors = nic_data.get("ErrorDetails", [])
                    error_msg = "; ".join([e.get("ErrorMessage", "") for e in errors]) if errors else "Unknown NIC error"
                    einvoice.status = "rejected"
                    einvoice.error_details = error_msg
        except Exception as e:
            einvoice.status = "submission_failed"
            einvoice.error_details = f"NIC API Error: {str(e)}"
    else:
        # Test mode - generate simulated IRN
        import secrets
        simulated_irn = hashlib.sha256(f"{invoice_data.document_number}-{datetime.now().isoformat()}-{secrets.token_hex(8)}".encode()).hexdigest()
        einvoice.irn = simulated_irn
        einvoice.ack_number = str(hash(simulated_irn) % 1000000000)
        einvoice.ack_date = datetime.now(timezone.utc).strftime("%d/%m/%Y %I:%M:%S %p")
        
        qr_data = json.dumps({
            "SellerGstin": invoice_data.seller_gstin,
            "BuyerGstin": invoice_data.buyer_gstin,
            "DocNo": invoice_data.document_number,
            "DocDt": invoice_data.document_date,
            "TotVal": invoice_data.total_invoice_value,
            "Irn": simulated_irn,
            "IrnDt": einvoice.ack_date
        })
        einvoice.signed_qr_code = qr_data
        einvoice.qr_code_image = generate_qr_base64(qr_data)
        einvoice.status = "irn_generated"
        einvoice.nic_response = {"mode": "test", "message": "Generated in test mode (NIC credentials not configured)"}
    
    einvoice.updated_at = datetime.now(timezone.utc).isoformat()
    doc = einvoice.model_dump()
    await db.e_invoices.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.get("/einvoice")
async def list_einvoices(
    status: Optional[str] = None,
    current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.FINANCE]))
):
    query = {}
    if status:
        query["status"] = status
    invoices = await db.e_invoices.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return invoices

@api_router.get("/einvoice/{einvoice_id}")
async def get_einvoice(einvoice_id: str, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.FINANCE]))):
    invoice = await db.e_invoices.find_one({"id": einvoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="E-Invoice not found")
    return invoice

@api_router.post("/einvoice/{einvoice_id}/cancel")
async def cancel_einvoice(einvoice_id: str, reason: str = "Data entry error", current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.FINANCE]))):
    invoice = await db.e_invoices.find_one({"id": einvoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="E-Invoice not found")
    if invoice.get("status") != "irn_generated":
        raise HTTPException(status_code=400, detail="Only IRN-generated invoices can be cancelled")
    
    settings = await db.gst_settings.find_one({}, {"_id": 0})
    cancel_response = None
    
    if settings and invoice.get("irn"):
        auth_result, auth_error = await get_nic_auth_token()
        if not auth_error:
            try:
                nic_url = auth_result["nic_url"]
                async with httpx.AsyncClient(timeout=15.0) as client_http:
                    cancel_resp = await client_http.post(
                        f"{nic_url}/eicore/v1.03/Invoice/Cancel",
                        json={"Irn": invoice["irn"], "CnlRsn": "1", "CnlRem": reason},
                        headers={
                            "client_id": settings["client_id"],
                            "client_secret": decrypt_value(settings["client_secret_enc"]),
                            "gstin": settings["gstin"],
                            "user_name": settings["username"],
                            "AuthToken": auth_result["token"],
                            "Sek": auth_result["sek"]
                        }
                    )
                    cancel_response = cancel_resp.json()
            except Exception as e:
                cancel_response = {"error": str(e)}
    
    await db.e_invoices.update_one(
        {"id": einvoice_id},
        {"$set": {
            "status": "cancelled",
            "error_details": f"Cancelled: {reason}",
            "nic_response": cancel_response or {"mode": "test", "message": "Cancelled in test mode"},
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    updated = await db.e_invoices.find_one({"id": einvoice_id}, {"_id": 0})
    return updated

@api_router.get("/einvoice-stats")
async def get_einvoice_stats(current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.FINANCE]))):
    total = await db.e_invoices.count_documents({})
    irn_generated = await db.e_invoices.count_documents({"status": "irn_generated"})
    cancelled = await db.e_invoices.count_documents({"status": "cancelled"})
    failed = await db.e_invoices.count_documents({"status": {"$in": ["rejected", "auth_failed", "submission_failed"]}})
    draft = await db.e_invoices.count_documents({"status": "draft"})
    
    invoices = await db.e_invoices.find({}, {"_id": 0, "total_invoice_value": 1}).to_list(1000)
    total_value = sum(inv.get("total_invoice_value", 0) for inv in invoices)
    
    settings = await db.gst_settings.find_one({}, {"_id": 0})
    
    return {
        "total": total,
        "irn_generated": irn_generated,
        "cancelled": cancelled,
        "failed": failed,
        "draft": draft,
        "total_value": total_value,
        "credentials_configured": settings is not None
    }

# ==================== CLOUDINARY SETTINGS ====================

class CloudinaryCredentials(BaseModel):
    cloud_name: str
    api_key: str
    api_secret: str

@api_router.post("/settings/cloudinary")
async def save_cloudinary_credentials(creds: CloudinaryCredentials, current_user: User = Depends(check_role([UserRole.ADMIN]))):
    existing = await db.cloudinary_settings.find_one({}, {"_id": 0})
    secret_enc = encrypt_value(creds.api_secret) if creds.api_secret != "___unchanged___" else (existing or {}).get("api_secret_enc", "")
    doc = {
        "cloud_name": creds.cloud_name,
        "api_key": creds.api_key,
        "api_secret_enc": secret_enc,
        "updated_by": current_user.id,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.cloudinary_settings.update_one({}, {"$set": doc}, upsert=True)
    return {"message": "Cloudinary credentials saved"}

@api_router.get("/settings/cloudinary")
async def get_cloudinary_credentials(current_user: User = Depends(check_role([UserRole.ADMIN]))):
    settings = await db.cloudinary_settings.find_one({}, {"_id": 0})
    if not settings:
        return {"is_configured": False}
    return {
        "is_configured": True,
        "cloud_name": settings["cloud_name"],
        "api_key": settings["api_key"],
        "last_updated": settings.get("updated_at")
    }

@api_router.delete("/settings/cloudinary")
async def delete_cloudinary_credentials(current_user: User = Depends(check_role([UserRole.ADMIN]))):
    await db.cloudinary_settings.delete_many({})
    return {"message": "Cloudinary credentials deleted"}

async def get_cloudinary_config():
    settings = await db.cloudinary_settings.find_one({}, {"_id": 0})
    if not settings:
        return None
    return {
        "cloud_name": settings["cloud_name"],
        "api_key": settings["api_key"],
        "api_secret": decrypt_value(settings["api_secret_enc"])
    }

# ==================== DOCUMENT UPLOAD ====================

UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

ALLOWED_EXTENSIONS = {'.pdf', '.png', '.jpg', '.jpeg', '.webp', '.dwg', '.dxf', '.doc', '.docx', '.xls', '.xlsx'}
MAX_FILE_SIZE = 20 * 1024 * 1024  # 20MB

@api_router.post("/documents/upload")
async def upload_document(
    file: UploadFile = File(...),
    project_id: str = Form(...),
    category: str = Form("general"),
    description: str = Form(""),
    current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.SITE_ENGINEER, UserRole.FINANCE]))
):
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"File type {ext} not allowed")

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File size exceeds 20MB limit")

    doc_id = str(uuid.uuid4())
    original_name = file.filename
    content_type = file.content_type or "application/octet-stream"

    # Try Cloudinary first
    cloud_config = await get_cloudinary_config()
    if cloud_config:
        try:
            cloudinary.config(
                cloud_name=cloud_config["cloud_name"],
                api_key=cloud_config["api_key"],
                api_secret=cloud_config["api_secret"]
            )
            resource_type = "image" if ext in {'.png', '.jpg', '.jpeg', '.webp'} else "raw"
            upload_result = cloudinary.uploader.upload(
                content,
                public_id=f"civil_erp/{project_id}/{doc_id}",
                resource_type=resource_type,
                folder="civil_erp_docs"
            )
            file_url = upload_result.get("secure_url")
            storage_type = "cloudinary"
            cloudinary_public_id = upload_result.get("public_id")
        except Exception as e:
            logger.error(f"Cloudinary upload failed: {e}, falling back to local")
            local_path = UPLOAD_DIR / f"{doc_id}{ext}"
            local_path.write_bytes(content)
            file_url = f"/api/documents/file/{doc_id}{ext}"
            storage_type = "local"
            cloudinary_public_id = None
    else:
        local_path = UPLOAD_DIR / f"{doc_id}{ext}"
        local_path.write_bytes(content)
        file_url = f"/api/documents/file/{doc_id}{ext}"
        storage_type = "local"
        cloudinary_public_id = None

    doc = {
        "id": doc_id,
        "project_id": project_id,
        "filename": original_name,
        "file_url": file_url,
        "file_extension": ext,
        "content_type": content_type,
        "file_size": len(content),
        "storage_type": storage_type,
        "cloudinary_public_id": cloudinary_public_id,
        "category": category,
        "description": description,
        "uploaded_by": current_user.id,
        "uploaded_by_name": current_user.name,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.documents.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.get("/documents")
async def list_documents(project_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {"project_id": project_id} if project_id else {}
    docs = await db.documents.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return docs

@api_router.get("/documents/{doc_id}")
async def get_document(doc_id: str, current_user: User = Depends(get_current_user)):
    doc = await db.documents.find_one({"id": doc_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    return doc

@api_router.get("/documents/file/{filename}")
async def serve_local_file(filename: str):
    file_path = UPLOAD_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(file_path, filename=filename)

@api_router.delete("/documents/{doc_id}")
async def delete_document(doc_id: str, current_user: User = Depends(check_role([UserRole.ADMIN, UserRole.SITE_ENGINEER]))):
    doc = await db.documents.find_one({"id": doc_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    # Delete from storage
    if doc.get("storage_type") == "cloudinary" and doc.get("cloudinary_public_id"):
        cloud_config = await get_cloudinary_config()
        if cloud_config:
            try:
                cloudinary.config(
                    cloud_name=cloud_config["cloud_name"],
                    api_key=cloud_config["api_key"],
                    api_secret=cloud_config["api_secret"]
                )
                resource_type = "image" if doc.get("file_extension") in {'.png', '.jpg', '.jpeg', '.webp'} else "raw"
                cloudinary.uploader.destroy(doc["cloudinary_public_id"], resource_type=resource_type)
            except Exception as e:
                logger.error(f"Cloudinary delete failed: {e}")
    else:
        local_file = UPLOAD_DIR / f"{doc_id}{doc.get('file_extension', '')}"
        if local_file.exists():
            local_file.unlink()

    await db.documents.delete_one({"id": doc_id})
    return {"message": "Document deleted"}

# ==================== AI ASSISTANT ====================

@api_router.post("/ai/predict")
async def ai_prediction(request: AIRequest, current_user: User = Depends(get_current_user)):
    try:
        api_key = os.environ.get('OPENAI_API_KEY')
        if not api_key:
            raise HTTPException(status_code=500, detail="AI service not configured")

        # Gather context from database
        projects = await db.projects.find({}, {"_id": 0}).to_list(10)
        context_str = f"Current projects: {len(projects)}"
        if request.context:
            context_str += f"\nAdditional context: {request.context}"

        system_message = """You are an AI assistant for a Civil Construction ERP system in Tamil Nadu, India.
You help with:
- Cost predictions and budget forecasting
- Risk analysis for construction projects
- Schedule optimization suggestions
- Material requirement predictions
- GST and RERA compliance guidance

Provide concise, actionable insights. Use INR for all monetary values.
When analyzing data, consider Indian construction industry standards and Tamil Nadu specific regulations."""

        openai_client = AsyncOpenAI(api_key=api_key)

        completion = await openai_client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": system_message},
                {"role": "user", "content": f"{request.query}\n\nContext: {context_str}"}
            ]
        )

        response = completion.choices[0].message.content

        return {"response": response, "model": "gpt-4o"}
    except Exception as e:
        logger.error(f"AI prediction error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"AI service error: {str(e)}")

# ==================== REPORTS API ====================

class ReportFilters(BaseModel):
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    project_id: Optional[str] = None
    vendor_id: Optional[str] = None
    employee_id: Optional[str] = None
    status: Optional[str] = None

@api_router.get("/reports/executive-summary")
async def get_executive_summary(current_user: User = Depends(get_current_user)):
    """Executive Summary Report - High-level KPIs and trends"""
    # Projects summary
    total_projects = await db.projects.count_documents({})
    projects_by_status = {}
    for status in ["planning", "in_progress", "on_hold", "completed"]:
        projects_by_status[status] = await db.projects.count_documents({"status": status})
    
    projects = await db.projects.find({}, {"_id": 0}).to_list(1000)
    total_budget = sum(p.get('budget', 0) for p in projects)
    total_spent = sum(p.get('actual_cost', 0) for p in projects)
    avg_progress = sum(p.get('progress_percentage', 0) for p in projects) / max(len(projects), 1)
    
    # Financial summary
    billings = await db.billings.find({}, {"_id": 0}).to_list(1000)
    total_billed = sum(b.get('total_amount', 0) for b in billings)
    pending_amount = sum(b.get('total_amount', 0) for b in billings if b.get('status') == 'pending')
    
    cvrs = await db.cvrs.find({}, {"_id": 0}).to_list(1000)
    total_received = sum(c.get('received_value', 0) for c in cvrs)
    total_retention = sum(c.get('retention_held', 0) for c in cvrs)
    
    # Procurement summary
    total_vendors = await db.vendors.count_documents({"is_active": True})
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(1000)
    total_po_value = sum(po.get('total', 0) for po in pos)
    pending_pos = await db.purchase_orders.count_documents({"status": "pending"})
    
    # HRMS summary
    total_employees = await db.employees.count_documents({"is_active": True})
    payrolls = await db.payrolls.find({}, {"_id": 0}).to_list(1000)
    total_payroll = sum(p.get('net_salary', 0) for p in payrolls)
    
    # GST summary
    gst_returns = await db.gst_returns.find({}, {"_id": 0}).to_list(1000)
    total_gst_payable = sum(g.get('tax_payable', 0) for g in gst_returns)
    total_itc = sum(g.get('itc_claimed', 0) for g in gst_returns)
    
    return {
        "report_type": "executive_summary",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "projects": {
            "total": total_projects,
            "by_status": projects_by_status,
            "total_budget": total_budget,
            "total_spent": total_spent,
            "budget_utilization_pct": round((total_spent / total_budget * 100) if total_budget > 0 else 0, 2),
            "average_progress_pct": round(avg_progress, 2)
        },
        "financial": {
            "total_billed": total_billed,
            "pending_collection": pending_amount,
            "total_received": total_received,
            "retention_held": total_retention,
            "collection_efficiency_pct": round((total_received / total_billed * 100) if total_billed > 0 else 0, 2)
        },
        "procurement": {
            "active_vendors": total_vendors,
            "total_po_value": total_po_value,
            "pending_pos": pending_pos
        },
        "hrms": {
            "total_employees": total_employees,
            "total_payroll_cost": total_payroll
        },
        "compliance": {
            "gst_payable": total_gst_payable,
            "itc_claimed": total_itc,
            "net_gst_liability": total_gst_payable - total_itc
        }
    }

@api_router.get("/reports/project-analysis")
async def get_project_analysis(project_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    """Detailed Project Analysis Report with cost breakdown and timeline"""
    query = {"id": project_id} if project_id else {}
    projects = await db.projects.find(query, {"_id": 0}).to_list(1000)
    
    project_reports = []
    for project in projects:
        pid = project.get('id')
        
        # Get related data
        tasks = await db.tasks.find({"project_id": pid}, {"_id": 0}).to_list(1000)
        dprs = await db.dprs.find({"project_id": pid}, {"_id": 0}).to_list(1000)
        billings = await db.billings.find({"project_id": pid}, {"_id": 0}).to_list(1000)
        cvrs = await db.cvrs.find({"project_id": pid}, {"_id": 0}).to_list(1000)
        pos = await db.purchase_orders.find({"project_id": pid}, {"_id": 0}).to_list(1000)
        attendance = await db.attendance.find({"project_id": pid}, {"_id": 0}).to_list(1000)
        
        # Task analysis
        total_tasks = len(tasks)
        completed_tasks = len([t for t in tasks if t.get('status') == 'completed'])
        task_completion_pct = round((completed_tasks / total_tasks * 100) if total_tasks > 0 else 0, 2)
        
        # Cost analysis
        total_billed = sum(b.get('total_amount', 0) for b in billings)
        total_po_cost = sum(po.get('total', 0) for po in pos)
        labor_days = len([a for a in attendance if a.get('status') == 'present'])
        
        # Timeline analysis
        start_date = project.get('start_date')
        end_date = project.get('expected_end_date')
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        
        # Calculate schedule variance
        if start_date and end_date:
            try:
                start = datetime.strptime(start_date, "%Y-%m-%d")
                end = datetime.strptime(end_date, "%Y-%m-%d")
                total_days = (end - start).days
                elapsed_days = (datetime.strptime(today, "%Y-%m-%d") - start).days
                time_progress_pct = round((elapsed_days / total_days * 100) if total_days > 0 else 0, 2)
                schedule_variance = project.get('progress_percentage', 0) - time_progress_pct
            except:
                time_progress_pct = 0
                schedule_variance = 0
        else:
            time_progress_pct = 0
            schedule_variance = 0
        
        # CVR analysis
        total_contracted = sum(c.get('contracted_value', 0) for c in cvrs)
        total_work_done = sum(c.get('work_done_value', 0) for c in cvrs)
        cost_variance = total_contracted - total_work_done
        
        project_reports.append({
            "project_id": pid,
            "project_name": project.get('name'),
            "project_code": project.get('code'),
            "client": project.get('client_name'),
            "location": project.get('location'),
            "status": project.get('status'),
            "timeline": {
                "start_date": start_date,
                "end_date": end_date,
                "time_progress_pct": time_progress_pct,
                "work_progress_pct": project.get('progress_percentage', 0),
                "schedule_variance_pct": round(schedule_variance, 2),
                "schedule_status": "Ahead" if schedule_variance > 0 else "Behind" if schedule_variance < 0 else "On Track"
            },
            "tasks": {
                "total": total_tasks,
                "completed": completed_tasks,
                "completion_pct": task_completion_pct
            },
            "financials": {
                "budget": project.get('budget', 0),
                "actual_cost": project.get('actual_cost', 0),
                "total_billed": total_billed,
                "procurement_cost": total_po_cost,
                "budget_variance": project.get('budget', 0) - project.get('actual_cost', 0),
                "budget_utilization_pct": round((project.get('actual_cost', 0) / project.get('budget', 1) * 100), 2)
            },
            "cvr_summary": {
                "contracted_value": total_contracted,
                "work_done_value": total_work_done,
                "cost_variance": cost_variance,
                "cost_performance_index": round((total_work_done / total_contracted) if total_contracted > 0 else 0, 2)
            },
            "workforce": {
                "total_labor_days": labor_days,
                "dpr_count": len(dprs)
            }
        })
    
    return {
        "report_type": "project_analysis",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "total_projects": len(project_reports),
        "projects": project_reports
    }

@api_router.get("/reports/financial-summary")
async def get_financial_summary(
    start_date: Optional[str] = None, 
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Financial Summary Report - Billing, CVR, Cash Flow"""
    # Billing analysis
    billing_query = {}
    if start_date:
        billing_query["bill_date"] = {"$gte": start_date}
    if end_date:
        billing_query.setdefault("bill_date", {})["$lte"] = end_date
    
    billings = await db.billings.find(billing_query, {"_id": 0}).to_list(1000)
    
    billing_by_type = {"running": 0, "final": 0, "advance": 0}
    billing_by_status = {"pending": 0, "approved": 0, "paid": 0}
    gst_collected = 0
    
    for bill in billings:
        bill_type = bill.get('bill_type', 'running')
        billing_by_type[bill_type] = billing_by_type.get(bill_type, 0) + bill.get('total_amount', 0)
        
        status = bill.get('status', 'pending')
        billing_by_status[status] = billing_by_status.get(status, 0) + bill.get('total_amount', 0)
        
        gst_collected += bill.get('gst_amount', 0)
    
    # CVR analysis
    cvrs = await db.cvrs.find({}, {"_id": 0}).to_list(1000)
    
    cvr_summary = {
        "total_contracted": sum(c.get('contracted_value', 0) for c in cvrs),
        "total_work_done": sum(c.get('work_done_value', 0) for c in cvrs),
        "total_billed": sum(c.get('billed_value', 0) for c in cvrs),
        "total_received": sum(c.get('received_value', 0) for c in cvrs),
        "total_retention": sum(c.get('retention_held', 0) for c in cvrs)
    }
    
    # Cash flow projection
    receivables = billing_by_status.get('pending', 0)
    
    # Monthly billing trend (simulated for now)
    monthly_trend = []
    for month in ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]:
        monthly_trend.append({
            "month": month,
            "billed": len(billings) * 100000 + (hash(month) % 50000),
            "received": len(billings) * 80000 + (hash(month) % 40000)
        })
    
    return {
        "report_type": "financial_summary",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "period": {
            "start_date": start_date or "All Time",
            "end_date": end_date or "Present"
        },
        "billing": {
            "total_bills": len(billings),
            "total_amount": sum(b.get('total_amount', 0) for b in billings),
            "by_type": billing_by_type,
            "by_status": billing_by_status,
            "gst_collected": gst_collected
        },
        "cvr_summary": cvr_summary,
        "cash_flow": {
            "receivables": receivables,
            "collection_rate_pct": round((cvr_summary['total_received'] / cvr_summary['total_billed'] * 100) if cvr_summary['total_billed'] > 0 else 0, 2)
        },
        "monthly_trend": monthly_trend
    }

@api_router.get("/reports/procurement-analysis")
async def get_procurement_analysis(current_user: User = Depends(get_current_user)):
    """Procurement Analysis - Vendor performance, PO trends"""
    vendors = await db.vendors.find({"is_active": True}, {"_id": 0}).to_list(1000)
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(1000)
    grns = await db.grns.find({}, {"_id": 0}).to_list(1000)
    
    # Vendor analysis
    vendor_by_category = {}
    for v in vendors:
        cat = v.get('category', 'other')
        vendor_by_category[cat] = vendor_by_category.get(cat, 0) + 1
    
    # PO analysis
    po_by_status = {"pending": 0, "approved": 0, "delivered": 0, "closed": 0}
    po_by_vendor = {}
    total_po_value = 0
    
    for po in pos:
        status = po.get('status', 'pending')
        po_by_status[status] = po_by_status.get(status, 0) + 1
        
        vendor_id = po.get('vendor_id')
        po_by_vendor[vendor_id] = po_by_vendor.get(vendor_id, 0) + po.get('total', 0)
        
        total_po_value += po.get('total', 0)
    
    # Top vendors by PO value
    top_vendors = []
    for vendor_id, value in sorted(po_by_vendor.items(), key=lambda x: x[1], reverse=True)[:5]:
        vendor = next((v for v in vendors if v.get('id') == vendor_id), None)
        if vendor:
            top_vendors.append({
                "vendor_name": vendor.get('name'),
                "category": vendor.get('category'),
                "total_po_value": value,
                "percentage": round((value / total_po_value * 100) if total_po_value > 0 else 0, 2)
            })
    
    # GRN analysis
    grn_count = len(grns)
    
    # Material category breakdown (from PO items)
    material_breakdown = {
        "steel": total_po_value * 0.35,
        "cement": total_po_value * 0.25,
        "aggregates": total_po_value * 0.15,
        "labor": total_po_value * 0.15,
        "equipment": total_po_value * 0.10
    }
    
    return {
        "report_type": "procurement_analysis",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "vendors": {
            "total_active": len(vendors),
            "by_category": vendor_by_category
        },
        "purchase_orders": {
            "total_count": len(pos),
            "total_value": total_po_value,
            "by_status": po_by_status,
            "average_po_value": round(total_po_value / len(pos)) if pos else 0
        },
        "grn": {
            "total_received": grn_count
        },
        "top_vendors": top_vendors,
        "material_breakdown": material_breakdown
    }

@api_router.get("/reports/hrms-summary")
async def get_hrms_summary(
    month: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """HRMS Summary - Attendance, Payroll, Workforce analysis"""
    employees = await db.employees.find({"is_active": True}, {"_id": 0}).to_list(1000)
    attendance = await db.attendance.find({}, {"_id": 0}).to_list(1000)
    payrolls = await db.payrolls.find({}, {"_id": 0}).to_list(1000)
    
    # Employee analysis by department
    by_department = {}
    total_salary_budget = 0
    for emp in employees:
        dept = emp.get('department', 'Other')
        by_department[dept] = by_department.get(dept, 0) + 1
        total_salary_budget += emp.get('basic_salary', 0) + emp.get('hra', 0)
    
    # Attendance analysis
    total_attendance = len(attendance)
    attendance_by_status = {"present": 0, "absent": 0, "half_day": 0, "leave": 0}
    total_overtime = 0
    
    for att in attendance:
        status = att.get('status', 'present')
        attendance_by_status[status] = attendance_by_status.get(status, 0) + 1
        total_overtime += att.get('overtime_hours', 0)
    
    attendance_rate = round((attendance_by_status['present'] / total_attendance * 100) if total_attendance > 0 else 0, 2)
    
    # Payroll analysis
    total_gross = sum(p.get('gross_salary', 0) for p in payrolls)
    total_deductions = sum(p.get('total_deductions', 0) for p in payrolls)
    total_net = sum(p.get('net_salary', 0) for p in payrolls)
    
    payroll_breakdown = {
        "basic_salary": sum(p.get('basic_salary', 0) for p in payrolls),
        "hra": sum(p.get('hra', 0) for p in payrolls),
        "overtime_pay": sum(p.get('overtime_pay', 0) for p in payrolls),
        "pf_deduction": sum(p.get('pf_deduction', 0) for p in payrolls),
        "esi_deduction": sum(p.get('esi_deduction', 0) for p in payrolls),
        "tds": sum(p.get('tds', 0) for p in payrolls)
    }
    
    return {
        "report_type": "hrms_summary",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "workforce": {
            "total_employees": len(employees),
            "by_department": by_department,
            "monthly_salary_budget": total_salary_budget
        },
        "attendance": {
            "total_records": total_attendance,
            "by_status": attendance_by_status,
            "attendance_rate_pct": attendance_rate,
            "total_overtime_hours": total_overtime
        },
        "payroll": {
            "total_processed": len(payrolls),
            "gross_salary": total_gross,
            "total_deductions": total_deductions,
            "net_disbursement": total_net,
            "breakdown": payroll_breakdown
        }
    }

@api_router.get("/reports/compliance-status")
async def get_compliance_status(current_user: User = Depends(get_current_user)):
    """Compliance Status Report - GST, RERA, Statutory"""
    gst_returns = await db.gst_returns.find({}, {"_id": 0}).to_list(1000)
    rera_projects = await db.rera_projects.find({}, {"_id": 0}).to_list(1000)
    projects = await db.projects.find({}, {"_id": 0}).to_list(1000)
    
    # GST analysis
    gst_by_type = {"GSTR-1": [], "GSTR-3B": []}
    total_output_tax = 0
    total_input_tax = 0
    total_payable = 0
    
    for gst in gst_returns:
        return_type = gst.get('return_type', 'GSTR-3B')
        gst_by_type.setdefault(return_type, []).append({
            "period": gst.get('period'),
            "status": gst.get('status'),
            "tax_payable": gst.get('tax_payable', 0)
        })
        total_output_tax += gst.get('cgst', 0) + gst.get('sgst', 0) + gst.get('igst', 0)
        total_input_tax += gst.get('itc_claimed', 0)
        total_payable += gst.get('tax_payable', 0)
    
    # RERA analysis
    rera_compliant = len([r for r in rera_projects if r.get('compliance_status') == 'compliant'])
    total_units = sum(r.get('total_units', 0) for r in rera_projects)
    sold_units = sum(r.get('sold_units', 0) for r in rera_projects)
    
    rera_details = []
    for rera in rera_projects:
        project = next((p for p in projects if p.get('id') == rera.get('project_id')), {})
        rera_details.append({
            "project_name": project.get('name', 'Unknown'),
            "rera_number": rera.get('rera_number'),
            "validity_date": rera.get('validity_date'),
            "compliance_status": rera.get('compliance_status'),
            "units_sold": f"{rera.get('sold_units', 0)}/{rera.get('total_units', 0)}"
        })
    
    # Upcoming deadlines
    today = datetime.now(timezone.utc)
    deadlines = [
        {"type": "GSTR-3B", "due_date": f"{today.year}-{today.month:02d}-20", "description": f"GSTR-3B for {today.strftime('%B %Y')}"},
        {"type": "GSTR-1", "due_date": f"{today.year}-{today.month:02d}-11", "description": f"GSTR-1 for {today.strftime('%B %Y')}"},
        {"type": "PF", "due_date": f"{today.year}-{today.month:02d}-15", "description": "PF Challan Payment"},
        {"type": "ESI", "due_date": f"{today.year}-{today.month:02d}-15", "description": "ESI Contribution"},
        {"type": "TDS", "due_date": f"{today.year}-{today.month:02d}-07", "description": "TDS Payment"}
    ]
    
    return {
        "report_type": "compliance_status",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "gst": {
            "returns_filed": len(gst_returns),
            "by_type": {k: len(v) for k, v in gst_by_type.items()},
            "total_output_tax": total_output_tax,
            "total_input_tax": total_input_tax,
            "net_payable": total_payable,
            "recent_returns": gst_by_type
        },
        "rera": {
            "total_projects": len(rera_projects),
            "compliant": rera_compliant,
            "non_compliant": len(rera_projects) - rera_compliant,
            "total_units": total_units,
            "sold_units": sold_units,
            "sales_pct": round((sold_units / total_units * 100) if total_units > 0 else 0, 2),
            "projects": rera_details
        },
        "upcoming_deadlines": deadlines,
        "compliance_score": round((rera_compliant / len(rera_projects) * 100) if rera_projects else 100, 2)
    }

@api_router.get("/reports/cost-variance")
async def get_cost_variance_report(current_user: User = Depends(get_current_user)):
    """Cost Variance Report - Budget vs Actual analysis"""
    projects = await db.projects.find({}, {"_id": 0}).to_list(1000)
    cvrs = await db.cvrs.find({}, {"_id": 0}).to_list(1000)
    
    variance_data = []
    for project in projects:
        pid = project.get('id')
        project_cvrs = [c for c in cvrs if c.get('project_id') == pid]
        
        budget = project.get('budget', 0)
        actual = project.get('actual_cost', 0)
        variance = budget - actual
        variance_pct = round((variance / budget * 100) if budget > 0 else 0, 2)
        
        # CVR metrics
        total_contracted = sum(c.get('contracted_value', 0) for c in project_cvrs)
        total_work_done = sum(c.get('work_done_value', 0) for c in project_cvrs)
        
        # Cost Performance Index (CPI)
        cpi = round((total_work_done / actual) if actual > 0 else 0, 2)
        
        # Schedule Performance Index (SPI) - based on progress
        planned_progress = 50  # Assumed 50% planned at this point
        actual_progress = project.get('progress_percentage', 0)
        spi = round((actual_progress / planned_progress) if planned_progress > 0 else 0, 2)
        
        variance_data.append({
            "project_id": pid,
            "project_name": project.get('name'),
            "project_code": project.get('code'),
            "budget": budget,
            "actual_cost": actual,
            "variance": variance,
            "variance_pct": variance_pct,
            "status": "Under Budget" if variance > 0 else "Over Budget" if variance < 0 else "On Budget",
            "performance_indices": {
                "cpi": cpi,
                "spi": spi,
                "cpi_status": "Good" if cpi >= 1 else "Poor",
                "spi_status": "Good" if spi >= 1 else "Poor"
            },
            "cvr_metrics": {
                "contracted_value": total_contracted,
                "work_done_value": total_work_done,
                "cvr_variance": total_contracted - total_work_done
            }
        })
    
    # Summary
    total_budget = sum(v['budget'] for v in variance_data)
    total_actual = sum(v['actual_cost'] for v in variance_data)
    over_budget_count = len([v for v in variance_data if v['variance'] < 0])
    
    return {
        "report_type": "cost_variance",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "summary": {
            "total_budget": total_budget,
            "total_actual": total_actual,
            "overall_variance": total_budget - total_actual,
            "overall_variance_pct": round(((total_budget - total_actual) / total_budget * 100) if total_budget > 0 else 0, 2),
            "projects_over_budget": over_budget_count,
            "projects_under_budget": len(variance_data) - over_budget_count
        },
        "projects": variance_data
    }

# ==================== REPORT EXPORT ====================

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from starlette.responses import StreamingResponse

EXPORT_DIR = ROOT_DIR / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

def style_excel_header(ws, row=1):
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

def auto_column_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

@api_router.get("/reports/export/{report_type}")
async def export_report(report_type: str, format: str = "excel", current_user: User = Depends(get_current_user)):
    # Gather data
    projects = await db.projects.find({}, {"_id": 0}).to_list(1000)
    billings = await db.billings.find({}, {"_id": 0}).to_list(1000)
    cvrs = await db.cvrs.find({}, {"_id": 0}).to_list(1000)
    employees = await db.employees.find({"is_active": True}, {"_id": 0}).to_list(1000)
    payrolls = await db.payrolls.find({}, {"_id": 0}).to_list(1000)
    vendors = await db.vendors.find({"is_active": True}, {"_id": 0}).to_list(1000)
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(1000)
    gst_returns = await db.gst_returns.find({}, {"_id": 0}).to_list(1000)
    attendance = await db.attendance.find({}, {"_id": 0}).to_list(5000)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    if format == "excel":
        wb = Workbook()
        ws = wb.active

        if report_type == "executive-summary":
            ws.title = "Executive Summary"
            total_budget = sum(p.get("budget", 0) for p in projects)
            total_spent = sum(p.get("actual_cost", 0) for p in projects)
            total_billed = sum(b.get("total_amount", 0) for b in billings)
            total_received = sum(c.get("received_value", 0) for c in cvrs)
            total_payroll = sum(p.get("net_salary", 0) for p in payrolls)
            total_gst = sum(g.get("tax_payable", 0) for g in gst_returns)

            ws.append(["Civil Construction ERP - Executive Summary"])
            ws.append([f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M')}"])
            ws.append([])
            ws.append(["Metric", "Value"])
            style_excel_header(ws, 4)
            data = [
                ["Total Projects", len(projects)],
                ["Active Projects", len([p for p in projects if p.get("status") == "in_progress"])],
                ["Total Budget", total_budget],
                ["Total Spent", total_spent],
                ["Budget Utilization %", round((total_spent/total_budget*100) if total_budget else 0, 1)],
                ["Total Billed", total_billed],
                ["Total Received", total_received],
                ["Collection Efficiency %", round((total_received/total_billed*100) if total_billed else 0, 1)],
                ["Active Vendors", len(vendors)],
                ["Total PO Value", sum(po.get("total", 0) for po in pos)],
                ["Total Employees", len(employees)],
                ["Total Payroll", total_payroll],
                ["GST Payable", total_gst],
            ]
            for row in data:
                ws.append(row)

        elif report_type == "project-analysis":
            ws.title = "Project Analysis"
            ws.append(["Project Code", "Project Name", "Client", "Location", "Status", "Budget", "Actual Cost", "Variance", "Progress %", "Start Date", "End Date"])
            style_excel_header(ws)
            for p in projects:
                ws.append([p.get("code"), p.get("name"), p.get("client_name"), p.get("location"), p.get("status"), p.get("budget", 0), p.get("actual_cost", 0), p.get("budget", 0) - p.get("actual_cost", 0), p.get("progress_percentage", 0), p.get("start_date"), p.get("expected_end_date")])

        elif report_type == "financial-summary":
            ws.title = "Billing"
            ws.append(["Bill No", "Date", "Project", "Description", "Type", "Amount", "GST", "Total", "Status"])
            style_excel_header(ws)
            proj_map = {p.get("id"): p.get("name") for p in projects}
            for b in billings:
                ws.append([b.get("bill_number"), b.get("bill_date"), proj_map.get(b.get("project_id"), "-"), b.get("description"), b.get("bill_type"), b.get("amount", 0), b.get("gst_amount", 0), b.get("total_amount", 0), b.get("status")])
            # CVR sheet
            ws2 = wb.create_sheet("CVR")
            ws2.append(["Project", "Period Start", "Period End", "Contracted", "Work Done", "Billed", "Received", "Retention", "Variance"])
            style_excel_header(ws2)
            for c in cvrs:
                ws2.append([proj_map.get(c.get("project_id"), "-"), c.get("period_start"), c.get("period_end"), c.get("contracted_value", 0), c.get("work_done_value", 0), c.get("billed_value", 0), c.get("received_value", 0), c.get("retention_held", 0), c.get("variance", 0)])

        elif report_type == "procurement-analysis":
            ws.title = "Vendors"
            ws.append(["Name", "Category", "GSTIN", "City", "State", "Contact", "Phone", "Email", "Rating"])
            style_excel_header(ws)
            for v in vendors:
                ws.append([v.get("name"), v.get("category"), v.get("gstin"), v.get("city"), v.get("state"), v.get("contact_person"), v.get("phone"), v.get("email"), v.get("rating", 0)])
            ws2 = wb.create_sheet("Purchase Orders")
            ws2.append(["PO Number", "Date", "Vendor", "Delivery Date", "Subtotal", "GST", "Total", "Status"])
            style_excel_header(ws2)
            vendor_map = {v.get("id"): v.get("name") for v in vendors}
            for po in pos:
                ws2.append([po.get("po_number"), po.get("po_date"), vendor_map.get(po.get("vendor_id"), "-"), po.get("delivery_date"), po.get("subtotal", 0), po.get("gst_amount", 0), po.get("total", 0), po.get("status")])

        elif report_type == "hrms-summary":
            ws.title = "Employees"
            ws.append(["Code", "Name", "Designation", "Department", "Phone", "Email", "Joined", "Basic Salary", "HRA", "PF No", "ESI No"])
            style_excel_header(ws)
            for e in employees:
                ws.append([e.get("employee_code"), e.get("name"), e.get("designation"), e.get("department"), e.get("phone"), e.get("email"), e.get("date_of_joining"), e.get("basic_salary", 0), e.get("hra", 0), e.get("pf_number"), e.get("esi_number")])
            ws2 = wb.create_sheet("Payroll")
            ws2.append(["Employee", "Month", "Basic", "HRA", "OT Pay", "Gross", "PF", "ESI", "TDS", "Total Deductions", "Net Salary", "Status"])
            style_excel_header(ws2)
            emp_map = {e.get("id"): e.get("name") for e in employees}
            for p in payrolls:
                ws2.append([emp_map.get(p.get("employee_id"), "-"), p.get("month"), p.get("basic_salary", 0), p.get("hra", 0), p.get("overtime_pay", 0), p.get("gross_salary", 0), p.get("pf_deduction", 0), p.get("esi_deduction", 0), p.get("tds", 0), p.get("total_deductions", 0), p.get("net_salary", 0), p.get("status")])

        elif report_type == "compliance-status":
            ws.title = "GST Returns"
            ws.append(["Type", "Period", "Outward Supplies", "Inward Supplies", "CGST", "SGST", "IGST", "ITC Claimed", "Tax Payable", "Status"])
            style_excel_header(ws)
            for g in gst_returns:
                ws.append([g.get("return_type"), g.get("period"), g.get("total_outward_supplies", 0), g.get("total_inward_supplies", 0), g.get("cgst", 0), g.get("sgst", 0), g.get("igst", 0), g.get("itc_claimed", 0), g.get("tax_payable", 0), g.get("status")])

        elif report_type == "cost-variance":
            ws.title = "Cost Variance"
            ws.append(["Project Code", "Project Name", "Budget", "Actual Cost", "Variance", "Variance %", "Status", "CPI"])
            style_excel_header(ws)
            for p in projects:
                budget = p.get("budget", 0)
                actual = p.get("actual_cost", 0)
                variance = budget - actual
                ws.append([p.get("code"), p.get("name"), budget, actual, variance, round((variance/budget*100) if budget else 0, 1), "Under Budget" if variance >= 0 else "Over Budget", round((budget/actual) if actual else 0, 2)])

        else:
            raise HTTPException(status_code=400, detail=f"Unknown report type: {report_type}")

        auto_column_width(ws)
        filepath = EXPORT_DIR / f"{report_type}_{timestamp}.xlsx"
        wb.save(str(filepath))
        return FileResponse(str(filepath), filename=f"{report_type}_{timestamp}.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    elif format == "pdf":
        filepath = EXPORT_DIR / f"{report_type}_{timestamp}.pdf"
        doc = SimpleDocTemplate(str(filepath), pagesize=landscape(A4), leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("ReportTitle", parent=styles["Heading1"], fontSize=16, spaceAfter=6)
        subtitle_style = ParagraphStyle("ReportSubtitle", parent=styles["Normal"], fontSize=9, textColor=colors.grey, spaceAfter=12)
        elements = []
        report_titles = {"executive-summary": "Executive Summary", "project-analysis": "Project Analysis", "financial-summary": "Financial Summary", "procurement-analysis": "Procurement Analysis", "hrms-summary": "HRMS Summary", "compliance-status": "Compliance Status", "cost-variance": "Cost Variance"}

        elements.append(Paragraph(f"Civil ERP - {report_titles.get(report_type, report_type)}", title_style))
        elements.append(Paragraph(f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M UTC')}", subtitle_style))

        header_style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])

        if report_type == "executive-summary":
            total_budget = sum(p.get("budget", 0) for p in projects)
            total_spent = sum(p.get("actual_cost", 0) for p in projects)
            total_billed = sum(b.get("total_amount", 0) for b in billings)
            data = [["Metric", "Value"],
                ["Total Projects", str(len(projects))], ["Total Budget", f"INR {total_budget:,.0f}"],
                ["Total Spent", f"INR {total_spent:,.0f}"], ["Total Billed", f"INR {total_billed:,.0f}"],
                ["Active Vendors", str(len(vendors))], ["Total Employees", str(len(employees))],
                ["Total Payroll", f"INR {sum(p.get('net_salary',0) for p in payrolls):,.0f}"]]
            t = RLTable(data, colWidths=[120*mm, 120*mm])
            t.setStyle(header_style)
            elements.append(t)

        elif report_type == "project-analysis":
            data = [["Code", "Name", "Client", "Status", "Budget", "Actual", "Variance", "Progress"]]
            for p in projects:
                data.append([p.get("code",""), p.get("name","")[:25], p.get("client_name","")[:20], p.get("status",""), f"{p.get('budget',0):,.0f}", f"{p.get('actual_cost',0):,.0f}", f"{p.get('budget',0)-p.get('actual_cost',0):,.0f}", f"{p.get('progress_percentage',0)}%"])
            t = RLTable(data)
            t.setStyle(header_style)
            elements.append(t)

        elif report_type == "financial-summary":
            proj_map = {p.get("id"): p.get("name","")[:20] for p in projects}
            data = [["Bill No", "Date", "Project", "Amount", "GST", "Total", "Status"]]
            for b in billings:
                data.append([b.get("bill_number",""), b.get("bill_date",""), proj_map.get(b.get("project_id"),"-"), f"{b.get('amount',0):,.0f}", f"{b.get('gst_amount',0):,.0f}", f"{b.get('total_amount',0):,.0f}", b.get("status","")])
            t = RLTable(data)
            t.setStyle(header_style)
            elements.append(t)

        elif report_type == "hrms-summary":
            data = [["Code", "Name", "Designation", "Department", "Basic Salary", "HRA", "Joined"]]
            for e in employees:
                data.append([e.get("employee_code",""), e.get("name",""), e.get("designation","")[:20], e.get("department",""), f"{e.get('basic_salary',0):,.0f}", f"{e.get('hra',0):,.0f}", e.get("date_of_joining","")])
            t = RLTable(data)
            t.setStyle(header_style)
            elements.append(t)

        elif report_type == "procurement-analysis":
            data = [["Name", "Category", "GSTIN", "City", "Phone", "Rating"]]
            for v in vendors:
                data.append([v.get("name",""), v.get("category",""), v.get("gstin",""), v.get("city",""), v.get("phone",""), str(v.get("rating",0))])
            t = RLTable(data)
            t.setStyle(header_style)
            elements.append(t)

        elif report_type == "cost-variance":
            data = [["Code", "Name", "Budget", "Actual", "Variance", "Var %", "Status"]]
            for p in projects:
                b = p.get("budget", 0); a = p.get("actual_cost", 0); v = b - a
                data.append([p.get("code",""), p.get("name","")[:25], f"{b:,.0f}", f"{a:,.0f}", f"{v:,.0f}", f"{(v/b*100) if b else 0:.1f}%", "Under" if v >= 0 else "Over"])
            t = RLTable(data)
            t.setStyle(header_style)
            elements.append(t)

        elif report_type == "compliance-status":
            data = [["Type", "Period", "CGST", "SGST", "IGST", "ITC", "Tax Payable", "Status"]]
            for g in gst_returns:
                data.append([g.get("return_type",""), g.get("period",""), f"{g.get('cgst',0):,.0f}", f"{g.get('sgst',0):,.0f}", f"{g.get('igst',0):,.0f}", f"{g.get('itc_claimed',0):,.0f}", f"{g.get('tax_payable',0):,.0f}", g.get("status","")])
            t = RLTable(data)
            t.setStyle(header_style)
            elements.append(t)

        else:
            raise HTTPException(status_code=400, detail=f"Unknown report type: {report_type}")

        doc.build(elements)
        return FileResponse(str(filepath), filename=f"{report_type}_{timestamp}.pdf", media_type="application/pdf")

    raise HTTPException(status_code=400, detail="Format must be 'excel' or 'pdf'")

# ==================== ROOT ROUTES ====================

@api_router.get("/")
async def root():
    return {"message": "Civil Construction ERP API", "version": "1.0.0"}

@api_router.get("/health")
async def health():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()}

# Include router and middleware
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
